import numpy as np
import math
import itertools
import os
from os import path
import json
import xlwings as xw
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows  # Correct import

from scipy.optimize import root, broyden1, newton_krylov

import warnings
from anchor_pro.report import EquipmentReport
import multiprocessing as mp
from functools import partial
import copy

from anchor_pro.utilities import Utilities
from anchor_pro.concrete_anchors import ConcreteCMU, ConcreteAnchors, CMUAnchors


def root_solver_worker(queue, residual_func, u_init, p, root_kwargs):
    try:
        res = root(residual_func, u_init, args=(p,), **root_kwargs)
        queue.put(res)
    except Exception as e:
        queue.put(e)

def solve_equilibrium_with_timeout(residual_func, u_init, p, timeout=5, **root_kwargs):
    queue = mp.Queue()
    process = mp.Process(target=root_solver_worker, args=(queue, residual_func, u_init, p, root_kwargs))
    process.start()
    process.join(timeout)

    if process.is_alive():
        process.terminate()
        process.join()
        return np.zeros_like(u_init), False

    if not queue.empty():
        result = queue.get()
        if isinstance(result, Exception):
            raise result
        return result.x, result.success
    else:
        print('Analysis timeout occurred')
        return np.zeros_like(u_init), False

class ExcelTablesImporter:
    NA_VALUES = ['NA', 'N/A', '-', 'None', ""]

    def __init__(self, excel_path):
        print(f'Reading data from Excel')
        self.path = excel_path
        wb = xw.Book(self.path)

        # Import Project Info from Named Cells on Project Worksheet
        project_sheet = wb.sheets['Project']
        self.project_info = {}
        for name in wb.names:
            if name.name.startswith('_xlfn') or name.name.startswith('_xlpm') or name.name.startswith('_xleta'):
                continue
            try:
                if name.refers_to_range.sheet == project_sheet:
                    value = name.refers_to_range.value
                    self.project_info[name.name] = value
            except Exception as e:
                print(f"Skipping {name.name} due to error: {e}")

        # Import Excel Tables
        self.df_equipment = None
        self.df_base_geometry = None
        self.df_wall_geometry = None
        self.df_concrete = None
        self.df_walls = None
        self.df_anchors = None
        self.df_brackets_catalog = None
        self.bracket_group_list = None
        self.df_product_groups = None

        # Table References (sheet name, table NW cell name, instance attribute name)
        table_references = [('Equipment', 'tbl_equipment', 'df_equipment'),
                            ('Base Geometry', 'tbl_base_geometry', 'df_base_geometry'),
                            ('Wall Geometry', 'tblBrackets', 'df_wall_geometry'),
                            ('Fastener Patterns', 'tblBacking', 'df_fasteners'),
                            ('Concrete', 'tbl_concrete', 'df_concrete'),
                            ('Walls', 'tblWalls', 'df_walls'),
                            ('Anchors', 'tbl_anchors', 'df_anchors'),
                            ('SMS', 'tblSMS', 'df_sms'),
                            ('Brackets', 'tblBracketCatalog', 'df_brackets_catalog'),
                            ('Anchor Product Groups', 'tblProductGroups', 'df_product_groups')]

        for (sheet_name, tbl_name, df_name) in table_references:
            sheet = wb.sheets[sheet_name]
            table_cell = sheet.range(tbl_name)
            start_address = table_cell.get_address(0, 0, include_sheetname=True, external=False)
            df = sheet.range(start_address).expand().options(pd.DataFrame,
                                                             header=1,
                                                             index=False,
                                                             expand='table').value
            setattr(self, df_name, df)

        # Replace undesired inputs
        self.df_equipment = self.df_equipment.replace(ExcelTablesImporter.NA_VALUES, None)

        # Adjust Indicies on Certain Tables
        for df in [self.df_walls, self.df_product_groups, self.df_fasteners]:
            df.set_index(df.columns[0], inplace=True)

        # Import Bracket Product Group Names
        self.bracket_group_list = wb.names['bracket_groups_list'].refers_to_range.value


class ProjectController:
    TABLE_COLUMNS = {
        # GENERAL INFO
        'Item': {'units': None, 'width': 24, 'alignment': 'l', 'style': None},
        'Group': {'units': None, 'width': 12, 'alignment': 'l', 'style': None},
        'Wp': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        'Fp': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        # BASE CONCRETE ANCHOR RESULTS
        'Base Anchor': {'units': None, 'width': 24, 'alignment': 'l', 'style': None},
        'Base Anchor Max Tension': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        'Base Thickness Check': {'units': None, 'width': None, 'alignment': None, 'style': 'condition'},
        'Base Spacing and Edge Checks': {'units': None, 'width': None, 'alignment': None, 'style': 'condition'},
        'Base Steel Tensile Strength': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Base Concrete Tension Breakout': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Base Anchor Pullout': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Base Side Face Blowout': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Base Bond Strength': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Base Steel Shear Strength': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Base Shear Breakout': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Base Shear Pryout': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Base Tension-Shear Interaction': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Base Anchor OK': {'units': None, 'width': None, 'alignment': None, 'style': 'result'},
        'Optimum Base Anchor': {'units': None, 'width': None, 'alignment': None, 'style': 'optimum'},
        # HARDWARE FASTENERS
        'Hardware SMS': {'units': None, 'width': None, 'alignment': 'l', 'style': None},
        'Hardware SMS Max Tension': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        'Hardware SMS Max Shear': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        'Hardware SMS Tension DCR': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Hardware SMS Shear DCR': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Hardware SMS OK': {'units': None, 'width': None, 'alignment': None, 'style': 'result'},
        'Optimum Hardware SMS': {'units': None, 'width': None, 'alignment': None, 'style': 'optimum'},
        # BASE STRAPS
        'Base Strap': {'units': None, 'width': 16, 'alignment': 'l', 'style': None},
        'Maximum Base Strap Tension': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        'Base Strap DCR': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Base Strap OK': {'units': None, 'width': None, 'alignment': None, 'style': 'result'},
        # WALL BRACKETS
        'Wall Bracket': {'units': None, 'width': 16, 'alignment': 'l', 'style': None},
        'Maximum Bracket Tension': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        'Bracket DCR': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Bracket OK': {'units': None, 'width': None, 'alignment': None, 'style': 'result'},
        # WALL SMS ANCHORS
        'Wall SMS': {'units': None, 'width': 8, 'alignment': 'l', 'style': None},
        'Wall SMS Max Tension': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        'Wall SMS Max Shear': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        'Wall SMS Tension DCR': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall SMS Shear DCR': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall SMS OK': {'units': None, 'width': None, 'alignment': None, 'style': 'result'},
        'Optimum Wall SMS': {'units': None, 'width': None, 'alignment': None, 'style': 'optimum'},
        # WALL CONCRETE ANCHOR
        'Wall Anchor': {'units': None, 'width': 24, 'alignment': 'l', 'style': None},
        'Wall Anchor Max Tension': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        'Wall Thickness Check': {'units': None, 'width': None, 'alignment': None, 'style': 'condition'},
        'Wall Spacing and Edge Checks': {'units': None, 'width': None, 'alignment': None, 'style': 'condition'},
        'Wall Steel Tensile Strength': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall Concrete Tension Breakout': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall Anchor Pullout': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall Side Face Blowout': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall Bond Strength': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall Steel Shear Strength': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall Shear Breakout': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall Shear Pryout': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall Tension-Shear Interaction': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall Anchor OK': {'units': None, 'width': None, 'alignment': None, 'style': 'result'},
        'Optimum Wall Anchor': {'units': None, 'width': None, 'alignment': None, 'style': 'optimum'}
    }

    def __init__(self, excel_path, output_dir):
        # Inputs
        self.excel_path = excel_path
        self.output_dir = output_dir
        self.excel_tables = None
        self.df_equipment = None  # Merge of equipment, base geometry, wall geometry, concrete tables
        self.model_inputs_dictionaries = {}
        self.get_model_inputs_from_excel_tables()

        # Multiprocessing pool
        self.pool = mp.Pool(processes=mp.cpu_count()) if self.excel_tables.project_info[
            'use_parallel_processing'] else None

        # Outputs
        self.items_for_report = {}
        self.governing_items = None
        self.group_dict = None
        self.results_table = None

    def input_validation(self):
        """ Function to validate all required user inputs are provided before attempting analysis"""
        pass

    @staticmethod
    def append_results(results_lists, model, is_optimum_base=False, is_optimum_wall=False):
        # Create blank row
        for key, results in results_lists.items():
            results.append('')

        # Populate item data
        results_lists['Item'][-1] = model.equipment_id
        results_lists['Group'][-1] = model.group
        results_lists['Wp'][-1] = model.Wp
        results_lists['Fp'][-1] = model.Fp

        # Populate base anchor data
        if model.base_anchors is not None:
            results_lists['Base Anchor'][-1] = model.base_anchors.anchor_id
            results_lists['Base Thickness Check'][-1] = \
                "OK" if model.base_anchors.spacing_requirements['slab_thickness_ok'] else "NG"
            results_lists['Base Spacing and Edge Checks'][-1] = \
                "OK" if model.base_anchors.spacing_requirements['edge_and_spacing_ok'] else "NG"

            for limit_state in ['Steel Tensile Strength',
                                'Concrete Tension Breakout',
                                'Anchor Pullout',
                                'Side Face Blowout',
                                'Bond Strength',
                                'Steel Shear Strength',
                                'Shear Pryout']:

                if model.omit_analysis:
                    result = 'NA'
                elif limit_state not in model.base_anchors.results.index:
                    result = 'NA'
                else:
                    result = model.base_anchors.results.loc[limit_state, 'Utilization']
                results_lists['Base ' + limit_state][-1] = result

            if model.omit_analysis:
                results_lists['Base Anchor Max Tension'][-1] = 'NA'
                results_lists['Base Shear Breakout'][-1] = 'NA'
            else:
                results_lists['Base Anchor Max Tension'][-1] = model.base_anchors.anchor_force_results[:, 0].max()
                cases = list(
                    set(model.base_anchors.shear_breakout_long_name_to_short_name_map.keys()) & set(
                        model.base_anchors.results.index))
                max_breakout = model.base_anchors.results.loc[cases, 'Utilization'].max()
                results_lists['Base Shear Breakout'][-1] = max_breakout

            results_lists['Base Tension-Shear Interaction'][
                -1] = 'NA' if model.base_anchors.DCR is None else model.base_anchors.DCR
            results_lists['Base Anchor OK'][
                -1] = False if model.base_anchors.DCR is None else (model.base_anchors.DCR < 1)

            results_lists['Optimum Base Anchor'][-1] = is_optimum_base

        # Populate Hardware Bracket Data  (Based on Governing Base Plate Connection or Wall Brace Connection
        floor_plate_sms_list = [plate.connection.anchors_obj for plate in model.floor_plates
                                if isinstance(plate.connection, SMSHardwareAttachment)
                                and plate.connection.anchors_obj.results]
        if floor_plate_sms_list:
            model.update_element_resultants(model.governing_solutions['base_anchor_tension']['sol'])
            floor_plate_sms = max(floor_plate_sms_list, key=lambda x: x.max_dcr())
            floor_plate_dcr = floor_plate_sms.max_dcr()
        else:
            floor_plate_sms = None
            floor_plate_dcr = -np.inf

        bracket_sms_list = [bracket.connection.anchors_obj for bracket in model.wall_brackets
                            if isinstance(bracket.connection, SMSHardwareAttachment)
                            and bracket.connection.anchors_obj.results]
        if bracket_sms_list:
            model.update_element_resultants(model.governing_solutions['wall_bracket_tension']['sol'])
            bracket_sms = max(bracket_sms_list, key=lambda x: x.max_dcr())
            bracket_sms_dcr = bracket_sms.max_dcr()
        else:
            bracket_sms = None
            bracket_sms_dcr = -np.inf

        if floor_plate_sms or bracket_sms:
            if floor_plate_dcr > bracket_sms_dcr:
                model.update_element_resultants(model.governing_solutions['base_anchor_tension']['sol'])
                governing_sms = floor_plate_sms
            else:
                governing_sms = bracket_sms

            results_lists['Hardware SMS'][-1] = governing_sms.screw_size
            results_lists['Hardware SMS Max Tension'][-1] = governing_sms.results['Tension Demand']
            results_lists['Hardware SMS Max Shear'][-1] = governing_sms.results['Shear Demand']
            results_lists['Hardware SMS Tension DCR'][-1] = governing_sms.results['Tension DCR']
            results_lists['Hardware SMS Shear DCR'][-1] = max(governing_sms.results['Shear X DCR'],
                                                              governing_sms.results['Shear Y DCR'])
            results_lists['Hardware SMS OK'][-1] = bool(governing_sms.results['OK'])

        # Populate Base Straps Data
        if model.base_straps and model.omit_analysis:
            results_lists['Base Strap'][-1] = model.base_straps[0].bracket_id
            results_lists['Maximum Base Strap Tension'][-1] = 'NA'
            results_lists['Base Strap DCR'][-1] = 'NA'
            results_lists['Base Strap OK'][-1] = 'NA'
        elif model.base_straps:
            governing_strap = max(model.base_straps, key=lambda x: x.tension_dcr)
            results_lists['Base Strap'][-1] = governing_strap.bracket_id
            results_lists['Maximum Base Strap Tension'][-1] = governing_strap.brace_force
            results_lists['Base Strap DCR'][-1] = governing_strap.tension_dcr
            results_lists['Base Strap OK'][-1] = bool(governing_strap.tension_dcr <= 1) if isinstance(governing_strap.tension_dcr, (int, float)) else governing_strap.tension_dcr

        # Populate Wall Bracket Data
        if model.wall_brackets:
            results_lists['Wall Bracket'][-1] = model.wall_brackets[0].bracket_id
            if not model.omit_analysis:
                idx_bracket, idx_theta = np.unravel_index(np.argmax(model.wall_bracket_forces[:, :, 0]),
                                                          model.wall_bracket_forces[:, :, 0].shape)
                results_lists['Maximum Bracket Tension'][-1] = np.max(
                    model.wall_bracket_forces[idx_bracket, idx_theta, 0])
                dcr = model.wall_brackets[idx_bracket].tension_dcr
                results_lists['Bracket DCR'][-1] = dcr
                results_lists['Bracket OK'][-1] = bool(dcr <= 1) if isinstance(dcr, (int, float)) else True

        # Populate Wall Anchor Data
        if model.wall_type == 'Metal Stud':
            #todo: Update for sms case to get governing anchor_obj from backing elements, not from equipment_model.wall_anchors

            # Determine wall_anchors object with maximum anchor tension
            walls_with_anchors = [(wall_anchor_obj, wall_anchor_obj.results['Tension DCR']) for
                                  wall_anchor_obj in model.wall_anchors.values() if
                                  wall_anchor_obj is not None and wall_anchor_obj.results]
            wall_anchors = max(walls_with_anchors, key=lambda x: x[1], default=(None, -np.inf))[0]

            if wall_anchors:
                results_lists['Wall SMS'][-1] = wall_anchors.screw_size
                if not model.omit_analysis:
                    results_lists['Wall SMS Max Tension'][-1] = wall_anchors.results['Tension Demand']
                    results_lists['Wall SMS Max Shear'][-1] = wall_anchors.results['Shear Demand']
                    results_lists['Wall SMS Tension DCR'][-1] = wall_anchors.results['Tension DCR']
                    results_lists['Wall SMS Shear DCR'][-1] = wall_anchors.results['Shear DCR']
                    results_lists['Wall SMS OK'][-1] = wall_anchors.results['OK']

                results_lists['Optimum Wall SMS'][-1] = is_optimum_wall
        elif model.wall_type == 'Concrete':
            # Determine wall_anchors object with maximum anchor tension
            walls_with_anchors = [(wall, wall.DCR) for wall in model.wall_anchors.values() if
                                  wall is not None]
            wall_anchors = max(walls_with_anchors, key=lambda x: x[1], default=(None, -np.inf))[0]

            if wall_anchors:
                results_lists['Wall Anchor'][-1] = wall_anchors.anchor_id
                results_lists['Wall Thickness Check'][-1] = \
                    "OK" if wall_anchors.spacing_requirements['slab_thickness_ok'] else "NG"
                results_lists['Wall Spacing and Edge Checks'][-1] = \
                    "OK" if wall_anchors.spacing_requirements['edge_and_spacing_ok'] else "NG"

                for limit_state in ['Steel Tensile Strength',
                                    'Concrete Tension Breakout',
                                    'Anchor Pullout',
                                    'Side Face Blowout',
                                    'Bond Strength',
                                    'Steel Shear Strength',
                                    'Shear Pryout']:

                    if model.omit_analysis:
                        result = 'NA'
                    elif limit_state not in wall_anchors.results.index:
                        result = 'NA'
                    else:
                        result = wall_anchors.results.loc[limit_state, 'Utilization']
                    results_lists['Wall ' + limit_state][-1] = result

                if model.omit_analysis:
                    results_lists['Wall Anchor Max Tension'][-1] = 'NA'
                    results_lists['Wall Shear Breakout'][-1] = 'NA'
                else:
                    results_lists['Wall Anchor Max Tension'][-1] = wall_anchors.anchor_force_results[:,
                                                                   0].max()
                    cases = list(
                        set(wall_anchors.shear_breakout_long_name_to_short_name_map.keys()) & set(wall_anchors.results.index))
                    max_breakout = wall_anchors.results.loc[cases, 'Utilization'].max()
                    results_lists['Wall Shear Breakout'][-1] = max_breakout

                results_lists['Wall Tension-Shear Interaction'][
                    -1] = 'NA' if wall_anchors.DCR is None else wall_anchors.DCR
                results_lists['Wall Anchor OK'][
                    -1] = False if wall_anchors.DCR is None else (wall_anchors.DCR < 1)

                results_lists['Optimum Wall Anchor'][-1] = is_optimum_wall

    def get_model_inputs_from_excel_tables(self):
        """Performs appropriate lookups across dataframes (equipment table, base geometry, etc.) and assembles
        a dictionary of data series representing the required inputs for each equipment model.

        the dictionary contains the following keys:
        'equipment_data'
        'concrete_data'
        'base_geometry'
        'anchor_geometry'

        """
        # Read Excel File and Import Data Tables as DataFrames
        self.excel_tables = ExcelTablesImporter(self.excel_path)

        # Merge Tables
        df_equipment = self.excel_tables.df_equipment
        df_equipment = df_equipment[df_equipment['omit_calc'] != True]

        # Merge Base Geometry
        df_equipment = df_equipment.merge(self.excel_tables.df_base_geometry,
                                          left_on='base_geometry', right_on='Pattern Name',
                                          how='left')
        df_equipment = df_equipment.drop(columns=['Pattern Name'])

        # Merge Wall Geometry
        df_equipment = df_equipment.merge(self.excel_tables.df_wall_geometry,
                                          left_on='wall_geometry', right_on='Pattern Name',
                                          how='left', suffixes=('_base', '_wall'))
        df_equipment = df_equipment.drop(columns=['Pattern Name'])

        # Merge Concrete Data
        # todo: [New Base Materials] merge additional tables for other material types
        df_equipment = df_equipment.merge(self.excel_tables.df_concrete,
                                          left_on='slab_id', right_on='slab_id',
                                          how='left', suffixes=('eq_', '_base'))
        df_equipment = df_equipment.merge(self.excel_tables.df_walls,
                                          left_on='wall_id', right_on='wall_id',
                                          how='left', suffixes=('_base', '_wall'))
        df_equipment = df_equipment.merge(self.excel_tables.df_concrete,
                                          left_on='slab_id_wall', right_on='slab_id',
                                          how='left', suffixes=('_base', '_wall'))

        ''' The decision is made deliberately not to merge anchor data or bracket data.
        These tables are independently quite large, and since checking by group may be selected,
        it seems preferable not to explicitly include these tables in df_equipment.'''

        self.df_equipment = df_equipment

    def run_analysis(self):

        df_results = pd.DataFrame()

        if self.excel_tables.project_info['use_parallel_processing']:
            # Parallel Processing
            print('\n')
            print('#' * 40)
            print('Beginning analysis (parallel processing)')

            items_and_results_list = self.pool.starmap(self.create_model_and_analyze_item,
                                                       [(self.excel_tables.project_info, equipment_data,
                                                         self.excel_tables.df_anchors,
                                                         self.excel_tables.df_product_groups,
                                                         self.excel_tables.df_brackets_catalog,
                                                         self.excel_tables.bracket_group_list,
                                                         self.excel_tables.df_fasteners,
                                                         self.excel_tables.df_sms) for index, equipment_data in
                                                        self.df_equipment.iterrows()])

            for (model, d) in items_and_results_list:
                self.items_for_report[model.equipment_id] = model
                df_results = pd.concat([df_results, pd.DataFrame(d)], ignore_index=True)

        else:
            # Serial Processing
            print('\n')
            print('#' * 40)
            print('Beginning analysis (serial processing)')
            print('#' * 40)

            results_list = []
            for index, equipment_data in self.df_equipment.iterrows():
                model, results = self.create_model_and_analyze_item(self.excel_tables.project_info, equipment_data,
                                                                    self.excel_tables.df_anchors,
                                                                    self.excel_tables.df_product_groups,
                                                                    self.excel_tables.df_brackets_catalog,
                                                                    self.excel_tables.bracket_group_list,
                                                                    self.excel_tables.df_fasteners,
                                                                    self.excel_tables.df_sms)
                results_list.append(results)
                self.items_for_report[model.equipment_id] = model

            for d in results_list:
                df_results = pd.concat([df_results, pd.DataFrame(d)], ignore_index=True)

        self.get_governing_by_group()
        print('Analysis Complete')

        print('\n')
        print('#' * 40)
        print('Compiling results')
        print('#' * 40)

        # Create a workbook and select the active worksheet
        self.df_results = df_results
        self.create_excel_summary_table(df_results)

    @staticmethod
    def create_model_and_analyze_item(project_info, equipment_data, anchor_catalog, anchor_product_groups,
                                      wall_bracket_catalog, bracket_groups_list, df_fasteners, df_sms):
        """ Function to run analysis of single model"""
        results_lists = {key: [] for key in ProjectController.TABLE_COLUMNS}

        # Create Model
        print(f'Creating {equipment_data["equipment_id"]} Model')
        model = EquipmentModel()
        model.set_model_data(project_info, equipment_data, df_fasteners, df_sms)
        model.calculate_fp()
        model.calculate_factored_loads()

        def get_anchor_product_list(group_name, member_type='Slab'):
            if group_name is None:
                return None

            anchor_products = anchor_product_groups[
                anchor_product_groups[group_name] == True].index
            filtered_anchors_df = anchor_catalog[anchor_catalog['product'].isin(anchor_products)]
            if member_type == 'Slab':
                filtered_anchors_df = filtered_anchors_df[filtered_anchors_df['slab_ok']]
            elif member_type == 'Filled Deck':
                filtered_anchors_df = filtered_anchors_df[filtered_anchors_df['deck_top_ok']]
            return filtered_anchors_df['anchor_id'].tolist()

        def get_sms_list(product_mode):
            if product_mode == 'Default':
                return ['No. 12']
            else:
                return ['No. 14', 'No. 12', 'No. 10', 'No. 8', 'No. 6']

        def get_bracket_group_list():
            idx = bracket_groups_list.index(model.bracket_group) + 1
            catalog_group = f'group_{idx}'
            filtered_bracket_df = wall_bracket_catalog[wall_bracket_catalog[catalog_group] == True]
            return filtered_bracket_df['bracket_id'].tolist()

        '''Collect Product Lists'''
        # Initialize Lists
        product_lists = {'base_anchor_list': [None],
                         'cxn_anchor_list': [None],
                         'bracket_list': [None],
                         'wall_anchor_list': [None]}

        product_applicable = {'base_anchor_list': model.base_anchors is not None,
                              'cxn_anchor_list': any([p.connection is not None for p in model.floor_plates]) or any(
                                  [b.connection is not None for b in model.wall_brackets]),
                              'bracket_list': len(model.wall_brackets) > 0,
                              'wall_anchor_list': any([v is not None for k, v in model.wall_anchors.items()]+
                                                      [b.anchors_obj is not None for b in model.wall_backing])}

        # Product List Parameters
        '''(list_object,
        material_applicable, 
        product_mode, 
        specified product, product group, 
        group_function, args, kwargs'''
        list_parameters = [
            # Base Concrete Anchors
            ('base_anchor_list',
             True,  # Replace with applicable to concrete when adding other base anchor types
             project_info['base_concrete_product_mode'],
             model.base_anchor_id, model.base_anchor_group,
             get_anchor_product_list, [model.base_anchor_group], {'member_type': model.profile_base}),
            # Hardware Connection SMS
            ('cxn_anchor_list',
             True,
             project_info['hardware_sms_product_mode'],
             model.cxn_sms_id, True,
             get_sms_list, [project_info['hardware_sms_product_mode']], {}),
            # Wall Brackets
            ('bracket_list',
             True,
             project_info['wall_bracket_product_mode'],
             model.bracket_id, model.bracket_group,
             get_bracket_group_list, [], {}),
            # Wall Concrete/CMU Anchors
            ('wall_anchor_list',
             model.wall_type in ['Concrete', 'CMU'],
             project_info['wall_concrete_product_mode'],
             model.wall_anchor_id, model.wall_anchor_group,
             get_anchor_product_list, [model.wall_anchor_group], {'member_type': 'Slab'}),
            # Wall SMS Anchors
            ('wall_anchor_list',
             model.wall_type == 'Metal Stud',
             project_info['wall_sms_product_mode'],
             model.wall_sms_id, True,
             get_sms_list, [project_info['wall_sms_product_mode']], {})]

        for list_name, material_applicable, mode, product, group, func, args, kwargs in list_parameters:
            if product_applicable[list_name] and material_applicable:
                if mode in ['Default', 'Specified Product'] and product:
                    product_lists[list_name] = [product]
                elif mode in ['Default', 'Product Group'] and group:
                    product_lists[list_name] = func(*args, **kwargs)

        # Verify That product is provided for all required items
        for list_name, product_list in product_lists.items():
            if product_applicable[list_name] and product_list == [None]:
                raise Exception(f"Must spcecify product or group for {model.equipment_id}, {list_name}")

        # Initialize Results Management Parameters
        optimum_base_cost = float('inf')
        optimum_wall_cost = float('inf')
        optimum_screw_index = 0
        optimum_base_results_index = None
        optimum_wall_results_index = None
        results_index = 0
        is_optimum_base = False
        is_optimum_wall = False
        item_for_report = None
        initial_solution_cache = None

        for base_anchor_id, bracket_id, wall_anchor_id, cxn_anchor_id in \
                itertools.product(product_lists['base_anchor_list'],
                                  product_lists['bracket_list'],
                                  product_lists['wall_anchor_list'],
                                  product_lists['cxn_anchor_list']):
            print(f'Analyzing {equipment_data["equipment_id"]} with '
                  f'base anchor: {base_anchor_id}, '
                  f'wall bracket: {bracket_id}, '
                  f'wall anchor: {wall_anchor_id}')

            base_anchor_data = None
            base_strap_data = None
            bracket_data = None
            wall_anchor_data = None

            if base_anchor_id:
                if model.base_material == 'Concrete':
                    base_anchor_data = anchor_catalog[anchor_catalog['anchor_id'] == base_anchor_id].iloc[0]
            if len(model.base_straps) > 0:
                base_strap_data = wall_bracket_catalog[wall_bracket_catalog['bracket_id'] == model.base_strap].iloc[0]
            if bracket_id:
                bracket_data = wall_bracket_catalog[wall_bracket_catalog['bracket_id'] == bracket_id].iloc[0]
            if wall_anchor_id:
                if model.wall_type in ['Concrete', 'CMU']:
                    wall_anchor_data = anchor_catalog[anchor_catalog['anchor_id'] == wall_anchor_id].iloc[0]
                elif model.wall_type == 'Metal Stud':
                    wall_anchor_data = wall_anchor_id  # For SMS, only "data" is anchor size

            model.set_product_data_and_analyze(base_anchor_data=base_anchor_data,
                                               base_strap_data=base_strap_data,
                                               bracket_data=bracket_data,
                                               wall_anchor_data=wall_anchor_data,
                                               hardware_screw_size=cxn_anchor_id,
                                               initial_solution_cache=initial_solution_cache)

            # Determine Optimum Base Anchor
            if model.installation_type in ['Base Anchored',
                                           'Wall Brackets'] and model.base_anchors is not None:
                current_cost = base_anchor_data['cost_rank']
                anchor_applicable = False if model.base_anchors.DCR is None else model.base_anchors.DCR < 1
                is_optimum_base = anchor_applicable and current_cost < optimum_base_cost
                if is_optimum_base:
                    # Update the results_lists to mark the previous optimum as not applicable
                    if optimum_base_results_index is not None:
                        results_lists['Optimum Base Anchor'][optimum_base_results_index] = False

                    # Update the current optimum
                    optimum_base_cost = current_cost
                    optimum_base_results_index = results_index
                    item_for_report = copy.deepcopy(model)  # Deep copy to preserve the state

            # Determine Optimum Wall Anchor
            if model.installation_type in ['Wall Brackets', 'Wall Mounted']:
                if model.wall_type == 'Metal Stud':
                    current_screw_index = product_lists['wall_anchor_list'].index(wall_anchor_id)
                    anchor_applicable = all(
                        [backing.anchors_obj.results['OK'] for backing in model.wall_backing if
                         backing.anchors_obj is not None and backing.anchors_obj.results])
                    is_optimum_wall = anchor_applicable and current_screw_index >= optimum_screw_index

                    if is_optimum_wall:
                        # Update the results_lists to mark the previous optimum as not applicable
                        if optimum_wall_results_index is not None:
                            results_lists['Optimum Wall SMS'][optimum_wall_results_index] = False

                        # Update the current optimum
                        optimum_screw_index = current_screw_index
                        optimum_wall_results_index = results_index
                        if model.base_anchors is None:
                            item_for_report = copy.deepcopy(model)  # Deep copy to preserve the state
                elif model.wall_type == 'Concrete':
                    current_cost = wall_anchor_data['cost_rank']
                    dcr_list = [anchors.DCR for anchors in model.wall_anchors.values() if anchors]
                    dcr_ok = [dcr < 1 for dcr in dcr_list if dcr]
                    anchor_applicable = all(dcr_ok) if dcr_list else False
                    is_optimum_wall = anchor_applicable and current_cost < optimum_wall_cost
                    if is_optimum_wall:
                        # Update the results_lists to mark the previous optimum as not applicable
                        if optimum_wall_results_index is not None:
                            results_lists['Optimum Wall Anchor'][optimum_wall_results_index] = False

                        # Update the current optimum
                        optimum_wall_cost = current_cost
                        optimum_results_index = results_index
                        item_for_report = copy.deepcopy(model)  # Deep copy to preserve the state
                # todo: select optimum base plate or bracket attachment sms
            if initial_solution_cache is None and not model.omit_analysis:
                initial_solution_cache = model.equilibrium_solutions
            results_index += 1
            ProjectController.append_results(results_lists, model, is_optimum_base=is_optimum_base,
                                             is_optimum_wall=is_optimum_wall)

        if item_for_report is None:
            item_for_report = copy.deepcopy(model)



        return item_for_report, results_lists

    def check_product_group(self):
        """Checks all anchor products in the specified product group and reports the "minimum" anchor"""

    def create_excel_summary_table(self, df_results):
        print('Creating summary table')
        # Export and Format Results Summary Table (using Openpyxl)
        file_path = os.path.join(self.output_dir, 'Results Summary.xlsx')

        wb = Workbook()
        ws = wb.active

        # Populate the sheet with DataFrame excel_tables
        for r in dataframe_to_rows(df_results, index=False, header=True):
            ws.append(r)

        # Predefine Fill Colors
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        red_fill = PatternFill(start_color="FF7276", end_color="FF7276", fill_type="solid")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        def style_dcr(c):
            if isinstance(c.value, (int, float)) and c.value > 1:
                c.fill = red_fill

        def style_condition(c):
            if c.value == 'NG':
                c.fill = red_fill
            elif c.value == 'OK':
                c.fill = green_fill

        def style_result(c):
            if c.value:
                c.fill = green_fill
            else:
                c.fill = gray_fill

        def style_optimum(c):
            if c.value:
                c.fill = green_fill
            else:
                c.fill = gray_fill

        def style_header(c):
            c.fill = header_fill
            c.font = Font(color="FFFFFF")  # White Font
            c.border = thin_border
            c.alignment = Alignment(horizontal='left', vertical='bottom', text_rotation=45)

        def style_general(c):
            # Only apply number formatting if it's a number, not a boolean
            if isinstance(c.value, (int, float)) and not isinstance(c.value, bool):
                c.number_format = '0.00'
            c.border = thin_border
            if c.value == 'NA' or c.value == "":
                c.fill = gray_fill

        style_functions = {'dcr': style_dcr,
                           'optimum': style_optimum,
                           'result': style_result,
                           'condition': style_condition}

        # Loop through cells in the first (header) row and apply style_header(cell)
        for cell in ws[1]:  # First row for headers
            style_header(cell)

        # Loop through columns and apply formatting
        for i, label in enumerate(df_results.columns):
            # Set column width
            width = ProjectController.TABLE_COLUMNS.get(label, {}).get('width', None)
            if width:
                ws.column_dimensions[get_column_letter(i + 1)].width = width

            # Apply styles to the rest of the column
            for row in ws.iter_rows(min_row=2, min_col=i + 1, max_col=i + 1):
                for cell in row:
                    alignment = ProjectController.TABLE_COLUMNS.get(label, {}).get('alignment', None)
                    style = ProjectController.TABLE_COLUMNS.get(label, {}).get('style', None)

                    if alignment == 'l':
                        cell.alignment = Alignment(horizontal='left', vertical='bottom')
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='bottom')

                    if style:
                        style_functions[style](cell)
                    style_general(cell)

        ws.auto_filter.ref = "A1:" + get_column_letter(ws.max_column) + "1"

        # Save the workbook
        wb.save(file_path)

    def get_governing_by_group(self):
        self.governing_items = {}
        if self.excel_tables.project_info['report_by_group'] == 'Worst-Case':
            # Create dictionary of {group: list of items in group}
            df = self.df_equipment.set_index(self.df_equipment.columns[0])

            self.group_dict = {group: list(item.index) for group, item in df.groupby('group')}
            self.group_dict['ungrouped'] = df[df['group'].isna()].index.tolist()

            # Create dictionary whose keys are governing items, and values are their corresponding group
            for group, eq_list in self.group_dict.items():
                if group == 'ungrouped':
                    for id in eq_list:
                        self.governing_items[id] = (None, 0)
                    continue

                # todo: [Input Validation]: check anchorage type (asure all equipment items are the same, maybe require same geometry)
                if self.df_equipment[self.df_equipment['equipment_id'] == eq_list[0]]['installation_type'].iloc[
                    0] == 'Base Anchored':
                    anchor_tensions = [self.items_for_report[eq_id].base_anchors.DCR for eq_id in eq_list]
                    anchor_tensions = [np.inf if val is None else val for val in
                                       anchor_tensions]  # if Tu_max is None, it is assumed that there was a failed spacing requirment, and this should be taken as the governing unit.
                    max_idx = np.argmax(anchor_tensions)
                    self.governing_items[eq_list[max_idx]] = (group, max_idx)
                elif self.df_equipment[self.df_equipment['equipment_id'] == eq_list[0]]['installation_type'].iloc[
                    0] in ['Wall Brackets', 'Wall Mounted']:
                    anchor_tensions = [
                        max([anchors.DCR for wall, anchors in self.items_for_report[eq_id].wall_anchors.items() if
                             anchors is not None] +
                            [backing.anchors_obj.DCR for backing in self.items_for_report[eq_id].wall_backing if
                             backing.anchors_obj is not None]) for eq_id in eq_list]
                    anchor_tensions = [np.inf if val is None else val for val in anchor_tensions]
                    max_idx = np.argmax(anchor_tensions)
                    self.governing_items[eq_list[max_idx]] = (group, max_idx)
                else:
                    raise Exception(
                        f'Installation type {self.df_equipment.loc[eq_list[0], "installation_type"]} for Equipment ID {eq_list[0]} not supported')
            self.governing_items = {
                k:self.governing_items[k]
                for k in self.items_for_report.keys()
                if k in self.governing_items}
        else:
            self.group_dict = None
            self.governing_items = {key: (None, 0) for key in self.items_for_report.keys()}

    def create_report(self):
        """ Creates a pdf report of EquipmentModel instances included in self.items_for_report"""

        report = EquipmentReport(self.excel_tables.project_info, self.items_for_report, self.governing_items,
                                 self.group_dict, pool=self.pool)
        # report.generate_pdf(self.output_dir)


class EquipmentModel:

    def __init__(self):
        '''User Input Parameters'''
        # Identification Parameters
        self.equipment_id = None
        self.group = None
        self.equipment_type = None

        # Code parameters
        self.code_edition = None
        self.factor_type = 'LRFD'  # todo: [USER-INPUT] read these values from project info input table

        # Global Geometry Parameters
        self.Wp = None  # Item Weight (lbs)
        self.Bx = None  # Item width (in X)
        self.By = None  # Item depth (in Y)
        self.H = None  # Item Height (in Z)
        self.zCG = None  # Height of center of gravity
        self.ex = None  # Eccentricity of COG from assumed origin
        self.ey = None  # Eccentricity of COG from assumed origin
        self.installation_type = None

        # Code Parameters for Seismic Force Calculations
        self.code_pars = {}
        self.include_overstrength = False

        # Hardware Attachement Parameters
        self.gauge_equip = None
        self.fy_equip = None
        self.cxn_sms_id = None

        # Base Attachment Parameters
        self.profile_base = None  # Profile of concrete member at base
        self.base_material = None

        self.anchor_pattern = None
        self.base_anchor_id = None
        self.base_anchor_group = None

        # Wall Attachment Parameters
        self.wall_type = None
        self.profile_wall = None  # Profile of concrete wall member
        self.bracket_group = None
        self.bracket_id = None
        self.wall_anchor_group = None
        self.wall_anchor_id = None
        self.wall_sms_id = None
        self.wall_offsets = {'X+': None, 'X-': None, 'Y+': None, 'Y-': None}

        '''CALCULATED PARAMETERS'''
        # Seismic force results
        self.Fp_code = None
        self.Fp_max = None
        self.Fp_min = None
        self.Fp_dynamic = None
        self.Fp = None

        # Factored load parameters
        self.Eh = None
        self.Emh = None
        self.Ev = None
        self.Fuh = None  # Ultimate Loads [LRFD]
        self.Fuv_max = None
        self.Fuv_min = None
        self.Fah = None  # Applied Loads [ASD]
        self.Fav_max = None
        self.Fav_min = None
        self.Fv_min = None  # The applicable Fv_min using LRFD or ASD
        self.Fv_max = None  # The applicable Fv_max using LRFD or ASD
        self.Fh = None  # The applicable Fh using LRFD or ASD
        self.asd_lrfd_ratio = None  # ASD/LRFD ratio

        # Model Attachment Elements
        self.floor_plates = []
        self.base_anchors = None
        self.base_straps = []
        self.base_strap = None
        self._num_base_anchors_in_tension = 0  # Used for stiffness update logic in solver

        self.wall_brackets = []
        self.wall_bracket_forces = None
        self.wall_backing = []
        self.backing_indices = {'X+': [], 'X-': [], 'Y+': [], 'Y-': []}
        self.wall_anchors = {'X+': None, 'X-': None, 'Y+': None, 'Y-': None}

        # Analysis Parameters
        self.omit_analysis = False
        self.model_unstable = False
        self.n_dof = None
        self.disp_dofs = []
        self.floor_plate_dofs = None  # DOF connectivity array for floor plate elements
        self.K = None
        self._u_previous = None
        self.converged = []
        # self.

        # Results
        self.theta_z = None  # List of angles at which horizontal load is applied
        self.equilibrium_solutions = None  # Array of equilibrium solutions for all angles theta_z
        self.sol = None
        self.theta_z_max = None
        self.governing_solutions = {'base_anchor_tension': {'sol': None, 'theta_z': None},
                                    'base_anchor_shear': {'sol': None, 'theta_z': None},
                                    'wall_bracket_tension': {'sol': None, 'theta_z': None},
                                    'wall_bracket_shear': {'sol': None, 'theta_z': None},
                                    'wall_anchor_tension': {'sol': None, 'theta_z': None},
                                    'wall_anchor_shear': {'sol': None, 'theta_z': None}}

        # Report Parameters
        self.frontmatter_file = None
        self.endmatter_file = None
        self.include_pull_test = False
        self.omit_bracket_output = False

    def set_model_data(self, project_info, equipment_data, df_fasteners=None, df_sms=None):
        """Populates instance attributes from input panda series.
        The input excel_tables is read from the user-input spreadsheet and parsed using the ProjectController object."""

        # Load Project Parameters
        self.code_edition = project_info['code_edition']

        # Load Equipment Parameters
        for key in vars(self).keys():
            if key in equipment_data.index:
                setattr(self, key, equipment_data.at[key])

        # Manual Corrections
        if not self.ex:
            self.ex = 0
        if not self.ey:
            self.ey = 0

        # Append full file path to front and end matter references:
        self.frontmatter_file = os.path.join(project_info['auxiliary_folder'],
                                             self.frontmatter_file) if self.frontmatter_file else None
        if self.frontmatter_file and not os.path.exists(self.frontmatter_file):
            raise Exception(f'Frontmatter file {self.frontmatter_file} not found.')

        self.endmatter_file = os.path.join(project_info['auxiliary_folder'],
                                           self.endmatter_file) if self.endmatter_file else None
        if self.endmatter_file and not os.path.exists(self.endmatter_file):
            raise Exception(f'Endmatter file {self.endmatter_file} not found.')

        # Load Code-Specific Parameters
        if self.code_edition == 'CBC 1998, 16B':
            self.code_pars['Ip'] = project_info['Ip']
            self.code_pars['Z'] = project_info['Z']
            self.code_pars['Cp'] = equipment_data['Cp']  # CBC 1998, 16B parameter
            self.code_pars['cp_amplification'] = equipment_data['cp_amplification']
            self.code_pars['cp_category'] = equipment_data['cp_category']
            self.code_pars['below_grade'] = equipment_data['below_grade']
            self.code_pars['grade_factor'] = 1 if not self.code_pars['below_grade'] else 2 / 3
            self.code_pars['Cp_eff'] = None
            self.include_overstrength = False
        elif self.code_edition == 'ASCE 7-16':
            self.code_pars['ap'] = equipment_data['ap']
            self.code_pars['Rp'] = equipment_data['Rp']
            self.code_pars['Ip'] = equipment_data['Ip']
            self.code_pars['sds'] = project_info['sds']
            self.code_pars['z'] = equipment_data['z']
            self.code_pars['h'] = equipment_data['building_height']
            self.code_pars['omega'] = equipment_data['omega']
            self.code_pars['use_dynamic'] = False  # todo: [Future Feature] add a use dynamic toggle in workbook
            self.code_pars['ai'] = None
            self.code_pars['Ax'] = None

        # Base Anchored and Wall Bracket Installation Type
        if self.installation_type in ['Wall Brackets', 'Base Anchored']:
            # Create Floor Plate Elements
            base_geometry_json = equipment_data['Pattern Definition_base']
            if not isinstance(base_geometry_json,str):
                print(f"WARNING: {self.equipment_id} base geometry {equipment_data['base_geometry']} not found.")
            else:
                E_base = equipment_data['E_base']
                poisson = equipment_data['poisson_base']
                self.floor_plate_elements_from_json(base_geometry_json, E_base, poisson, df_fasteners, df_sms)

                # Create Base Anchor Element
                xy_anchors_list = [plate.xy_anchors for plate in self.floor_plates if plate.xy_anchors.size > 0]

                if xy_anchors_list:  # Ensure list is not empty before concatenation
                    xy_anchors = np.concatenate(xy_anchors_list, axis=0)
                else:
                    xy_anchors = np.array([])

                if len(xy_anchors) > 0:
                    if self.base_material == 'Concrete':
                        self.include_overstrength = True  # todo: [Calc Refinement] Define when overstrength can be omitted
                        self.base_anchors = ConcreteAnchors(equipment_data, xy_anchors, base_or_wall='base')
                    elif self.base_material == 'Wood':
                        pass
                        # self.base_anchors == WoodAnchors() # todo: [WOOD] Finish this
        # Wall Bracket and Wall Mounted Installation Type
        if self.installation_type in ['Wall Brackets', 'Wall Mounted']:
            # Create Wall Bracket Elements
            self.wall_bracket_backing_anchors_from_json(equipment_data, df_fasteners, df_sms)

            for position, wall_anchors in self.wall_anchors.items():
                if isinstance(wall_anchors, ConcreteCMU):
                    wall_anchors.set_data(equipment_data=equipment_data,
                                          xy_anchors=wall_anchors.xy_anchors,
                                          base_or_wall='wall')

    def floor_plate_elements_from_json(self, json_string, E_base, poisson, df_fasteners, df_sms):
        """Creates floor plate elements based on a json string provided by the user-input excel form"""
        orientations = {'X+': 0.0, 'X-': np.pi, 'Y+': np.pi / 2, 'Y-': 3 * np.pi / 2}
        pattern = json.loads(json_string)

        selected_straps = {item['element']['straps']['strap'] for item in pattern if item['element']['straps']['strap'] != 'null'}
        if len(selected_straps) == 0:
            self.base_strap = None
        elif len(selected_straps) > 1:
            print(
                'Warning: user should not specify different strap types for base plates in a single geometry pattern.')
            self.base_strap = selected_straps.pop()
        else:
            self.base_strap = selected_straps.pop()

        for item in pattern:
            element = item['element']
            shape = element['shape']
            anchors = element['anchors']
            straps = element['straps']
            layout = item['layout']
            shape_points = [[float(x) + float(xr) * self.Bx, float(y) + float(yr) * self.By] for x, y, xr, yr in
                            zip(shape['X'], shape['Y'], shape['Xr'], shape['Yr'])]
            anchor_points = [[float(x) + float(xr) * self.Bx, float(y) + float(yr) * self.By] for x, y, xr, yr in
                             zip(anchors['X'], anchors['Y'], anchors['Xr'], anchors['Yr'])]
            strap_geometry = [[float(x), float(y), float(z), float(dx), float(dy), float(dz)] for x, y, z, dx, dy, dz in
                              zip(straps['X'], straps['Y'], straps['Z'], straps['DX'], straps['DY'], straps['DZ']) if
                              self.base_strap]

            # Determine if releases are present
            release_keys = ['release_mx', 'release_my', 'release_mz',
                            'release_xp', 'release_xn', 'release_yp', 'release_yn', 'release_zp', 'release_zn']

            # Check if any value is True for the specified keys
            releases_present = any(value for key in release_keys for value in layout.get(key, []))

            if releases_present or element['check_fasteners']:
                # Create a separate floor_plate_element for each insertion in the layout
                for x, y, xr, yr, rotation_angle, reflection_angle, \
                    r_mx, r_my, r_mz, r_xp, r_xn, r_yp, r_yn, r_zp, r_zn in zip(
                    layout["X"], layout["Y"], layout["Xr"], layout["Yr"],
                    layout["Rotation"], layout["Reflection"],
                    layout["release_mx"], layout["release_my"], layout["release_mz"],
                    layout["release_xp"], layout["release_xn"],
                    layout["release_yp"], layout["release_yn"],
                    layout["release_zp"], layout["release_zn"]
                ):

                    translation = np.array([x + xr * self.Bx,
                                            y + yr * self.By])

                    boundary = Utilities.transform_points(shape_points, translation, rotation_angle, reflection_angle)

                    anchors = Utilities.transform_points(anchor_points, translation, rotation_angle, reflection_angle)

                    (x0, y0) = Utilities.transform_points([[element['x0'], element['y0']]],
                                                          translation, rotation_angle, reflection_angle)[0]
                    z0 = element['z0']

                    (xc, yc) = Utilities.transform_points([[element['xc'], element['yc']]],
                                                          translation, rotation_angle, reflection_angle)[0]
                    zc = element['zc']

                    self.floor_plates.append(FloorPlateElement([boundary],
                                                               xy_anchors=None if anchors == [] else anchors,
                                                               x0=x0,
                                                               y0=y0,
                                                               z0=z0,
                                                               xc=xc,
                                                               yc=yc,
                                                               zc=zc,
                                                               release_mx=r_mx,
                                                               release_my=r_my,
                                                               release_mz=r_mz,
                                                               release_xp=r_xp,
                                                               release_xn=r_xn,
                                                               release_yp=r_yp,
                                                               release_yn=r_yn,
                                                               release_zp=r_zp,
                                                               release_zn=r_zn,
                                                               E_base=E_base,
                                                               poisson=poisson))

                    plate = self.floor_plates[-1]
                    if element['check_fasteners']:
                        # get all the inputs sorted out
                        fastener_pattern = element['fastener_geometry_name']
                        data = df_fasteners.loc[fastener_pattern]
                        faying_local_angle = orientations[element['fastener_orientation']]
                        faying_local_vector = (np.cos(faying_local_angle), np.sin(faying_local_angle))
                        faying_global_vector = Utilities.transform_vectors([faying_local_vector],
                                                                           rotation_angle, reflection_angle)[0]
                        if np.isclose(faying_global_vector[0], 0):
                            B = self.By
                        elif np.isclose(faying_global_vector[1], 0):
                            B = self.Bx
                        else:
                            B = 0
                        w = data['W'] + data['Wr'] * B
                        h = data['H'] + data['Hr'] * self.H
                        if w == 0:
                            raise Exception(
                                "For base plate attachments with fasteners, you must specify aboslute dimension for fastener pattern or place floor plates orthogonal to reference box")

                        L_horiz = w - 2 * data['X Edge']
                        L_vert = h - 2 * data['Y Edge']
                        y_offset = data['Y Offset']
                        x_offset = data['X Offset']
                        place_by_horiz = data['X Placement']
                        place_by_vert = data['Y Placement']

                        xy_points = Utilities.compute_backing_xy_points(data['X Number'],
                                                                        data['Y Number'],
                                                                        L_horiz, L_vert, x_offset, y_offset,
                                                                        place_by_horiz=place_by_horiz,
                                                                        place_by_vert=place_by_vert)

                        plate.connection = SMSHardwareAttachment(w, h, xy_points, df_sms,
                                                                 centroid=(xc, yc, zc),
                                                                 normal_vector=(*faying_global_vector, 0),
                                                                 x_offset=x_offset,
                                                                 y_offset=y_offset)
                        gauge = self.gauge_equip if self.gauge_equip is not None else 18
                        fy = self.fy_equip if self.fy_equip is not None else 33
                        plate.connection.anchors_obj.set_sms_properties(gauge, fy)

                    for strap_pts in strap_geometry:
                        (x_eq, y_eq,) = Utilities.transform_points([[strap_pts[0] + strap_pts[3],
                                                                     strap_pts[1] + strap_pts[4]]],
                                                                   translation, rotation_angle, reflection_angle)[0]
                        z_eq = strap_pts[2] + strap_pts[5]

                        (x_pl, y_pl,) = Utilities.transform_points([[strap_pts[0],
                                                                     strap_pts[1]]],
                                                                   translation, rotation_angle, reflection_angle)[0]
                        z_pl = strap_pts[2]

                        self.base_straps.append(BaseStrap((x_eq, y_eq, z_eq), (x_pl, y_pl, z_pl), plate))


            else:
                bearing_boundaries = []
                xy_anchors = []
                for x, y, xr, yr, rotation_angle, reflection_angle in zip(layout['X'],
                                                                          layout['Y'],
                                                                          layout['Xr'],
                                                                          layout['Yr'],
                                                                          layout['Rotation'],
                                                                          layout['Reflection']):
                    translation = np.array([x + xr * self.Bx,
                                            y + yr * self.By])

                    boundary = Utilities.transform_points(shape_points, translation, rotation_angle, reflection_angle)
                    anchors = Utilities.transform_points(anchor_points, translation, rotation_angle, reflection_angle)

                    bearing_boundaries.append(boundary)
                    xy_anchors += anchors.tolist()

                xy_anchors = None if xy_anchors == [] else np.array(xy_anchors)
                self.floor_plates.append(FloorPlateElement(bearing_boundaries,
                                                           xy_anchors=xy_anchors,
                                                           x0=0.0,
                                                           y0=0.0,
                                                           z0=element['z0'],
                                                           xc=0.0,
                                                           yc=0.0,
                                                           zc=element['z0'],
                                                           E_base=E_base,
                                                           poisson=poisson))

    def wall_bracket_backing_anchors_from_json(self, equipment_data, df_fasteners, df_sms):

        wall_normal_vectors = {'X+': (-1, 0, 0),
                               'X-': (1, 0, 0),
                               'Y+': (0, -1, 0),
                               'Y-': (0, 1, 0)}

        attachment_normal_vectors = {'X+': (1, 0, 0),
                                     'X-': (-1, 0, 0),
                                     'Y+': (0, 1, 0),
                                     'Y-': (0, -1, 0),
                                     'Z+': (0, 0, 1),
                                     'Z-': (0, 0, -1)}

        b_dimension = {'X+': self.By,
                       'X-': self.By,
                       'Y+': self.Bx,
                       'Y-': self.Bx}

        wall_geometry_json = equipment_data['Pattern Definition_wall']
        if not isinstance(wall_geometry_json,str):
            print(f"WARNING: {self.equipment_id} wall geometry {equipment_data['wall_geometry']} not found.")
            return

        json_data = json.loads(wall_geometry_json)

        self.omit_bracket_output = json_data['omit_bracket_output']

        # Extract Wall Offsets
        for wall_key, offset_val in json_data['wall_offsets'].items():
            self.wall_offsets[wall_key] = offset_val

        # Extract bracket_locations to DataFrame
        bracket_locations = json_data['bracket_locations']
        df_bracket_locations = pd.DataFrame(bracket_locations)
        df_backing_groups = pd.DataFrame.from_dict(json_data['backing_groups'], orient='columns')
        df_brackets = df_bracket_locations.merge(df_backing_groups, left_on='Backing Group', right_on='group_number',
                                                 how='left')
        df_brackets = df_brackets.merge(df_fasteners, left_on='backing_pattern', right_on='Pattern Name', how='left')

        # Wall Properties
        L = equipment_data['wall_height']
        E = equipment_data['wall_E']
        I = equipment_data['wall_I']

        bracket_backing_map = {}
        for index, bracket in df_brackets.iterrows():
            # Attachment Point
            x0 = bracket['X'] + bracket['Xr'] * self.Bx
            y0 = bracket['Y'] + bracket['Yr'] * self.By
            z0 = bracket['Z'] + bracket['Zr'] * self.H
            xyz_0 = (x0, y0, z0)

            # Wall Normal Vector
            supporting_wall = bracket['Supporting Wall']
            normal_vec = wall_normal_vectors[supporting_wall]

            # Calculate Bracket Centerline Point with Connection Offset
            attachment_normal_vec = np.array([attachment_normal_vectors[supporting_wall]])
            cxn_offset = json_data['attachment_offset']
            x_offset, y_offset, z_offset = (attachment_normal_vec * cxn_offset + np.array(xyz_0))[0]
            xyz_offset = (x_offset, y_offset, z_offset)

            wall_gap = 0 if not self.wall_offsets[supporting_wall] else self.wall_offsets[supporting_wall]
            backing_depth = bracket['D']
            if supporting_wall == 'X+':
                xyz_f = (0.5 * self.Bx + (wall_gap - backing_depth), y_offset, z_offset)
            elif supporting_wall == 'X-':
                xyz_f = (-0.5 * self.Bx - (wall_gap - backing_depth), y_offset, z_offset)
            elif supporting_wall == 'Y+':
                xyz_f = (x_offset, 0.5 * self.By + (wall_gap - backing_depth), z_offset)
            elif supporting_wall == 'Y-':
                xyz_f = (x_offset, -0.5 * self.By - (wall_gap - backing_depth), z_offset)
            else:
                raise Exception('Supporting Wall Incorrectly Defined')

            # Stiffness Releases
            releases = (bracket['N+'], bracket['N-'], bracket['P+'], bracket['P-'], bracket['Z+'], bracket['Z-'])

            # Wall Flexibility
            a = z0
            b = L - a
            wall_flexibility = (a ** 2 + b ** 2) / (3 * E * I * L)  # Wall idealized as simple span

            self.wall_brackets.append(
                WallBracketElement(supporting_wall, xyz_0, xyz_f, normal_vec, wall_flexibility,
                                   releases=releases, xyz_offset=xyz_offset, e_cxn=cxn_offset))

            # Create Bracket Connection Elements
            if json_data['check_fasteners']:
                bracket_obj = self.wall_brackets[-1]
                fastener_pattern = json_data["fastener_geometry"]
                data = df_fasteners.loc[fastener_pattern]
                try:
                    faying_global_vector = attachment_normal_vectors[bracket['Attachment Normal']]
                except:
                    raise Exception("Must specify wall bracket attachment normal direction.")

                B = 0  # Note, wall brackets are set to ignore any relative dimension
                w = data['W'] + data['Wr'] * B
                h = data['H'] + data['Hr'] * self.H
                if np.isclose(w, 0.0) or np.isclose(h, 0.0):
                    raise Exception(
                        "For wall bracket attachments, you must specify absolute dimensions")

                L_horiz = w - 2 * data['X Edge']
                L_vert = h - 2 * data['Y Edge']
                y_offset = data['Y Offset']
                x_offset = data['X Offset']
                place_by_horiz = data['X Placement']
                place_by_vert = data['Y Placement']

                xy_points = Utilities.compute_backing_xy_points(data['X Number'],
                                                                data['Y Number'],
                                                                L_horiz, L_vert, x_offset, y_offset,
                                                                place_by_horiz=place_by_horiz,
                                                                place_by_vert=place_by_vert)

                bracket_obj.connection = SMSHardwareAttachment(w, h, xy_points, df_sms,
                                                               centroid=(x0, y0, z0),
                                                               normal_vector=faying_global_vector)
                gauge = self.gauge_equip if self.gauge_equip is not None else 18
                fy = self.fy_equip if self.fy_equip is not None else 33
                bracket_obj.connection.anchors_obj.set_sms_properties(gauge, fy)

            # Map bracket index to backing group
            bracket_backing_map[index] = bracket['Backing Group']

        # Create Bracket-Backing mapping array
        backing_bracket_map = {}
        for i, bracket in enumerate(self.wall_brackets):
            backing_group = bracket_backing_map[i]
            if backing_group not in backing_bracket_map:
                backing_bracket_map[backing_group] = []
            backing_bracket_map[backing_group].append(i)

        # Create Backing Elements
        for backing_group, bracket_indices in backing_bracket_map.items():
            if backing_group == 0:
                for i in bracket_indices:
                    backing_data = df_brackets.iloc[i]
                    self.create_backing_element(backing_data, [i], b_dimension)
            else:
                backing_data = df_brackets[df_brackets['Backing Group'] == backing_group].iloc[0]
                self.create_backing_element(backing_data, bracket_indices, b_dimension)

        # Create Wall-Backing Mapping Array
        for index, backing in enumerate(self.wall_backing):
            self.backing_indices[backing.supporting_wall].append(index)

        # Create Wall Anchor Elements
        supporting_walls = set([el.supporting_wall for el in self.wall_backing])
        wall_type = equipment_data['wall_type']
        if wall_type in ['Concrete', 'CMU']:
            for wall_position in supporting_walls:
                # Collect anchor points from all wall anchors
                xy_anchors = np.concatenate(
                    [plate.pz_anchors + plate.centroid for plate in
                     [self.wall_backing[i] for i in self.backing_indices[wall_position]]], axis=0)

                self.include_overstrength = True  # todo: [Calc Refinement] Define when overstrength can be omitted
                self.wall_anchors[wall_position] = ConcreteAnchors(
                    xy_anchors=xy_anchors) if wall_type == 'Concrete' else CMUAnchors(xy_anchors=xy_anchors)

        elif wall_type == 'Metal Stud':
            wall_data = equipment_data.loc[['stud_gauge', 'stud_fy', 'num_gyp']]
            num_gyp = wall_data['num_gyp']
            for backing in self.wall_backing:
                if num_gyp == 2:
                    condition_x = 'Condition 3'
                    condition_y = 'Condition 3'
                elif num_gyp == 1:
                    condition_x = 'Condition 2'
                    condition_y = 'Condition 2'
                else:
                    condition_x = 'Condition 1'
                    condition_y = 'Condition 1'

                if backing.backing_type == 'Wall Backing (Strut)' and backing.w > backing.h:
                    condition_y = 'Condition 4'  # todo: [Refine SMS Prying] could add check to verify if shear is perpendicular to axis of strut.
                elif backing.backing_type == 'Wall Backing (Strut)' and backing.h > backing.w:
                    condition_x = 'Condition 4'

                backing.anchors_obj = SMSAnchors(wall_data=wall_data, xy_anchors=backing.pz_anchors,
                                                              backing_type=backing.backing_type, df_sms=df_sms,
                                                              condition_x=condition_x, condition_y=condition_y)
        else:
            raise Exception(f'Wall type {wall_type} for {self.equipment_id} not supported')

    def create_backing_element(self, backing_data, bracket_indices, b_dimension_dict):
        supporting_wall = backing_data['Supporting Wall']
        backing_type = backing_data['Connection Type']
        b = b_dimension_dict[supporting_wall]  # Width of unit parallel to supporting wall
        wb = backing_data['W'] + backing_data['Wr'] * b  # Width of bracket
        hb = backing_data['H'] + backing_data['Hr'] * self.H  # Height of bracket
        db = backing_data['D']
        L_horiz = wb - 2 * backing_data['X Edge']
        L_vert = hb - 2 * backing_data['Y Edge']
        y_offset = backing_data['Y Offset']
        x_offset = backing_data['X Offset']
        place_by_horiz = backing_data['X Placement']
        place_by_vert = backing_data['Y Placement']

        xy_anchor_points = Utilities.compute_backing_xy_points(backing_data['X Number'], backing_data['Y Number'],
                                                               L_horiz, L_vert, x_offset, y_offset,
                                                               place_by_horiz=place_by_horiz,
                                                               place_by_vert=place_by_vert)

        # Convert Bracket locations from global XYZ coordinates to local NPZ Coordinates
        npz_brackets = np.array(
            [bracket.G @ bracket.xyz_0 for bracket in [self.wall_brackets[i] for i in bracket_indices]])
        pz_brackets = npz_brackets[:, [1, 2]]
        pz_cent = np.mean(pz_brackets, axis=0)
        self.wall_backing.append(
            WallBackingElement(wb, hb, db, xy_anchor_points, pz_brackets, bracket_indices, supporting_wall,
                               backing_type=backing_type, centroid=pz_cent))

    def number_degrees_of_freedom(self):
        """Numbers the degrees of freedom, including additional DOFs for floor plates with moment releases."""
        self.floor_plate_dofs = np.full((len(self.floor_plates), 6), [0, 1, 2, 3, 4, 5], dtype=int)
        dof_count = 6
        self.disp_dofs = [1, 1, 1, 0, 0, 0]
        for i, element in enumerate(self.floor_plates):
            if element.release_zp and element.release_zn:
                self.floor_plate_dofs[i, 2] = dof_count
                self.disp_dofs.append(1)
                dof_count += 1
            if element.release_mx: # or element.prying_mx:
                self.floor_plate_dofs[i, 3] = dof_count
                self.disp_dofs.append(0)
                dof_count += 1
            if element.release_my: # or element.prying_my:
                self.floor_plate_dofs[i, 4] = dof_count
                self.disp_dofs.append(0)
                dof_count += 1
            if element.release_mz:
                self.floor_plate_dofs[i, 5] = dof_count
                self.disp_dofs.append(0)
                dof_count += 1
        self.n_dof = dof_count

    def set_element_dofs(self):
        """Updates element constraint matrices and dof mapping based on global dofs"""
        for i, element in enumerate(self.floor_plates):
            element.set_dof_constraints(self.n_dof, self.floor_plate_dofs[i, :])
        for element in self.wall_brackets:
            element.set_dof_constraints(self.n_dof)

    # def check_base_plate_prying(self):
        # for each plate, identify maximum anchor force
        # update element resultants
        # check prying
        # store prying result for report (largest force, if no prying, any if yes prying)
        # if any have prying,

            # refactor model to add dofs at plates
            # reanalyze model.

    def set_base_anchor_data(self, anchor_data):
        """Sets anchor properties for anchor object and updates stiffness properties for floor plate elements"""
        self.base_anchors.set_mechanical_anchor_properties(anchor_data)
        for element in self.floor_plates:
            element.set_anchor_properties(self.base_anchors)

    def set_wall_bracket_data(self, bracket_data):
        for element in self.wall_brackets:
            element.set_bracket_properties(bracket_data)

    def set_base_strap_data(self, strap_data):
        for element in self.base_straps:
            element.set_brace_properties(strap_data)
            element.pre_compute_matrices()

    def calculate_fp(self, use_dynamic=False):
        if self.code_edition == 'ASCE 7-16':
            sds = self.code_pars['sds']
            Ip = self.code_pars['Ip']
            ap = self.code_pars['ap']
            Rp = self.code_pars['Rp']
            z = self.code_pars['z']
            h = self.code_pars['h']
            omega = self.code_pars['omega']
            Wp = self.Wp

            self.Fp_min = 0.3 * sds * Ip * Wp  # ASCE7-16 13.3-2
            self.Fp_max = 1.6 * sds * Ip * Wp  # ASCE7-16 13.3-3
            self.Fp_code = (0.4 * ap * sds * Wp * Ip / Rp) * (1 + 2 * z / h)  # ASCE7-16 13.3-1

            if use_dynamic:
                ai = self.code_pars['ai']
                Ax = self.code_pars['Ax']
                self.Fp_dynamic = (ap * Wp * Ip / Rp) * ai * Ax  # ASCE7-16 13.3-4
                self.Fp = max(min(self.Fp_max, self.Fp_dynamic), self.Fp_min)  # ASCE7-16 13.3.1.1
            else:
                self.Fp = max(min(self.Fp_max, self.Fp_code), self.Fp_min)  # ASCE7-16 13.3.1.1

            self.Eh = self.Fp  # ASCE7-16 13.3.1.1
            self.Emh = omega * self.Eh
            self.Ev = 0.2 * sds * Wp  # ASCE7-16 13.3.1.2

        elif self.code_edition == 'ASCE 7-22':
            raise Exception('Definition of fp for ASCE7-22 not yet defined')

        elif self.code_edition == 'CBC 1998, 16B':
            Z = self.code_pars['Z']
            Ip = self.code_pars['Ip']
            Cp = self.code_pars['Cp']
            max_amplification = {1: 999,
                                 2: 2,
                                 4: 3}
            Cp_eff = min([Cp * self.code_pars['cp_amplification'],
                          max_amplification[self.code_pars['cp_amplification']]]) * self.code_pars['grade_factor']
            self.code_pars['Cp_eff'] = Cp_eff
            Wp = self.Wp

            self.Fp = Z * Ip * Cp_eff * Wp
            self.Eh = self.Fp/0.7  # ASCE7-16 13.3.1.1
            self.Ev = self.Eh / 3
            self.Emh = self.Fp/0.7

        else:
            raise Exception('Specified code year not supported.')

    def calculate_factored_loads(self):
        # LRFD
        self.Fuv_min = -0.9 * self.Wp + 1.0 * self.Ev  # Minimum downward force
        self.Fuv_max = -1.2 * self.Wp - 1.0 * self.Ev  # Maximum downward force
        if self.include_overstrength:
            self.Fuh = 1.0 * self.Emh
        else:
            self.Fuh = 1.0 * self.Eh

        # ASD
        self.Fav_min = -0.6 * self.Wp + 0.7 * self.Ev
        self.Fav_max = -1.0 * self.Wp - 0.7 * self.Ev
        if self.include_overstrength:
            self.Fah = 0.7 * self.Emh
        else:
            self.Fah = 0.7 * self.Eh

        if self.factor_type == "LRFD":
            self.Fv_min = self.Fuv_min
            self.Fv_max = self.Fuv_max
            self.Fh = self.Fuh
        elif self.factor_type == 'ASD':
            self.Fv_min = self.Fav_min
            self.Fv_max = self.Fav_max
            self.Fh = self.Fah

        self.asd_lrfd_ratio = self.Fah / self.Fuh

    def set_product_data_and_analyze(self, base_anchor_data=None, base_strap_data=None,
                                     bracket_data=None, wall_anchor_data=None,
                                     hardware_screw_size=None,
                                     initial_solution_cache=None):

        self.omit_analysis = False

        # Set Base Anchor Data
        if self.base_anchors is not None:
            self.base_anchors.reset_results()
            if base_anchor_data is not None:
                self.set_base_anchor_data(base_anchor_data)
                self.base_anchors.check_anchor_spacing()
                if not all(self.base_anchors.spacing_requirements.values()):
                    self.omit_analysis = True

        # Set Bast Strap Data
        if base_strap_data is not None:
            self.set_base_strap_data(base_strap_data)

        # Set Wall Brackets and Wall Anchors Data
        if self.wall_brackets != [] and bracket_data is not None:
            self.set_wall_bracket_data(bracket_data)

            for position, wall_anchors in self.wall_anchors.items():
                if wall_anchors is not None and wall_anchor_data is not None:
                    wall_anchors.reset_results()
                    if isinstance(wall_anchors, ConcreteCMU):
                        wall_anchors.set_mechanical_anchor_properties(wall_anchor_data)
                        wall_anchors.check_anchor_spacing()
                        if not all(wall_anchors.spacing_requirements.values()):
                            self.omit_analysis = True

            for wall_anchors in [b.anchors_obj for b in self.wall_backing]:
                if isinstance(wall_anchors, SMSAnchors) and wall_anchor_data is not None:
                    wall_anchors.reset_results()
                    wall_anchors.set_screw_size(wall_anchor_data)

        # Set Hardware Attachment Screw Size
        for plate in self.floor_plates:
            if isinstance(plate.connection, SMSHardwareAttachment) and hardware_screw_size is not None:
                plate.connection.anchors_obj.reset_results()
                plate.connection.anchors_obj.set_screw_size(hardware_screw_size)
        for bracket in self.wall_brackets:
            if isinstance(bracket.connection, SMSHardwareAttachment) and hardware_screw_size is not None:
                bracket.connection.anchors_obj.reset_results()
                bracket.connection.anchors_obj.set_screw_size(hardware_screw_size)

        # Number DOFs and Set-up Element Constraints
        self.number_degrees_of_freedom()
        self.set_element_dofs()

        # Check Model Stability
        self.check_model_stability()

        if not self.omit_analysis:
            self.analyze_model(initial_solution_cache=initial_solution_cache)

            # Base Anchor Checks
            if self.base_anchors is not None:
                self.base_anchors.get_governing_anchor_group()
                self.base_anchors.check_anchor_capacities()

            # Base Plate Connection Checks
            # for plate in self.floor_plates:
            #     if isinstance(plate.connection, SMSHardwareAttachment) and hardware_screw_size is not None:
            #         plate.connection.anchors_obj.check_anchors(self.asd_lrfd_ratio)

            # Base Strap Checks
            for strap in self.base_straps:
                asd_lrfd_ratio = None if strap.capacity_method == 'LRFD' else self.asd_lrfd_ratio
                strap.check_brace(self.governing_solutions['base_anchor_tension']['sol'], asd_lrfd_ratio=asd_lrfd_ratio)

            # Wall Bracket Checks
            if self.wall_brackets:
                bracket_max_tension = self.wall_bracket_forces[:, :, 0].max(axis=1)
                bracket = max([(b, t) for b, t in zip(self.wall_brackets, bracket_max_tension)], key=lambda x: x[1])[0]
                bracket.check_brackets()

            # Bracket Connection Checks
            # for bracket in self.wall_brackets:
            #     if isinstance(bracket.connection, SMSHardwareAttachment) and hardware_screw_size is not None:
            #         bracket.connection.anchors_obj.check_anchors(self.asd_lrfd_ratio)

            # Wall Anchor Checks
            all_wall_anchors = [anchors_obj for wall_position, anchors_obj in self.wall_anchors.items()
                                if anchors_obj is not None] + [b.anchors_obj for b in self.wall_backing if
                                                               b.anchors_obj is not None]
            for wall_anchors in all_wall_anchors:
                if isinstance(wall_anchors, SMSAnchors):
                    wall_anchors.check_anchors(self.asd_lrfd_ratio)
                elif isinstance(wall_anchors, CMUAnchors):
                    pass
                elif isinstance(wall_anchors, ConcreteAnchors):
                    wall_anchors.get_governing_anchor_group()
                    wall_anchors.check_anchor_capacities()
                elif wall_anchors is None:
                    pass
                else:
                    raise Exception("Anchor type not supported")

    def get_load_vector(self, theta_z):
        """ Return a load vector in the form: [Vx, Vy, N, Mx, My, T].
        Assumes self.calculate_factored_loads has been previously called."""
        #  Loads at basic DOFs
        vx = self.Fh * math.cos(theta_z)
        vy = self.Fh * math.sin(theta_z)
        if self.installation_type in ['Base Anchored', "Wall Brackets"]:
            fz = self.Fv_min
        else:
            fz = self.Fv_max

        mx = -vy * self.zCG + fz * self.ey
        my = vx * self.zCG - fz * self.ex
        t = -vx * self.ey + vy * self.ex

        # Zero applied moment at additional floor-plate DOFs
        p = np.zeros(self.n_dof)

        # Load Vector
        p[0:6] = np.array([vx, vy, fz, mx, my, t])

        return p

    def check_model_stability(self, tol=1e-12):
        """ Can be run after setting all anchor and hardware data.
        Will impose small dof displacements in principal directions
        and verify model stabilit by checking for non-zero eigenvalues."""

        dof_labels = ['Dx', 'Dy', 'Dz', 'Rx', 'Ry', 'Rz']
        dir_labels = ['+', '-']

        u0 = 1 * np.eye(self.n_dof)
        for dof, u_dof in enumerate(u0[0:5, :]):
            for dir, u in enumerate([u_dof, -u_dof]):
                k = self.update_stiffness_matrix(u)
                # eigvals = np.linalg.eigvals(k)
                # zero_modes = np.sum(np.abs(eigvals) < tol)
                # unstable = zero_modes > 1
                p = k @ u
                unstable = np.abs(p[dof]) < tol

                if unstable:
                    if self.installation_type == 'Wall Brackets' and dof == 2 and dir == 0:
                        '''Ignore Dz+ instability for Wall Brackets units.
                        It is assumed that resultant vertical forces will always be downward.'''
                        continue
                    else:
                        self.model_unstable = self.omit_analysis = True
                        print(f'WARNING: Model instability detected at DOF {dof_labels[dof] + dir_labels[dir]}. '
                              f'Check model geometry definitions, material properties, and releases.')
                        return

    def get_initial_dof_guess(self, theta_z):
        """ Returns an initial guess for the DOF displacements U0 = [dx, dy, dz, rx, ry, rz, ...]"""

        # u0 = self._p_proportional_dof_guess(theta_z)

        u_init_list = [self._p_proportional_dof_guess(theta_z),  # Proportional to p at given angle
                       self._p_proportional_dof_guess(theta_z + np.pi / 16),  # Proportional to p at perturbed angle
                       self._p_proportional_dof_guess(theta_z - np.pi / 16)]  # Proportional to p at perturbed angle
        # np.array([0, 0, 1e-6, 0, 0, 0] + [0]*(self.n_dof-6)),  # Uplift Translation Only
        # np.array([0, 0, -1e-6, 0, 0, 0] + [0]*(self.n_dof-6)),  # Gravity Translation Only
        # np.zeros(self.n_dof)]  # Zero

        return u_init_list

    def get_interpolated_dof_guess(self, idx):
        """ Returns trial solutions for an unconverged point
        by taking weighted averages of bounding converged solutions points"""

        n = len(self.converged)
        before = None
        after = None

        # Wrapped search backward
        for offset in range(1, n):
            i = (idx - offset) % n
            if self.converged[i]:
                sol_before = self.equilibrium_solutions[:, i]
                break

        # Wrapped search forward
        for offset in range(1, n):
            i = (idx + offset) % n
            if self.converged[i]:
                sol_after = self.equilibrium_solutions[:, i]
                break

        # sol_before = self.equilibrium_solutions[:, before]
        # sol_after = self.equilibrium_solutions[:, after]

        u_tries = [0.5 * sol_before + 0.5 * sol_after,
                   0.25 * sol_before + 0.75 * sol_after,
                   0.75 * sol_before + 0.25 * sol_after,
                   sol_after]

        return u_tries

    def _initial_stiffness_dof_guess_UNUSED(self, theta):
        p = self.get_load_vector(theta)
        u = np.linalg.solve(self.initial_stiffness_matrix_UNUSED(), p)
        return u

    def _p_proportional_dof_guess(self, theta):
        """ Determines the load vector corresponding to theta and returns an initial dof array proportional to p"""
        p = self.get_load_vector(theta)
        u = np.zeros(self.n_dof)
        p_unit_disp = p[0:3] if p[0:3].sum() == 0 else p[0:3] / np.linalg.norm(p[0:3])
        p_unit_rot = p[3:6] if p[0:6].sum() == 0 else p[3:6] / np.linalg.norm(p[3:6])
        dx, dy, dz = 1e-6 * p_unit_disp

        '''todo: Refine initial uplift guess by considering aspect ratio and ratio of overturning to restoring moment.
         Heuristically, item centroid must uplift for tipping about edge. Centerline is set to uplift'''
        delta_x = 0.5*self.Bx-self.ex if (theta<=np.pi/2 or theta >= 3*np.pi/2) else 0.5*self.Bx+self.ex
        delta_y = 0.5*self.By-self.ey if 0<=theta<=np.pi else 0.5*self.By+self.ey
        l_from_x = np.inf if np.isclose(np.cos(theta), 0) else delta_x / np.cos(theta)
        l_from_y = np.inf if np.isclose(np.sin(theta), 0) else delta_y / np.sin(theta)
        l_ot_approx = min(abs(l_from_x),abs(l_from_y))
        m_ot = (p[3]**2+p[4]**2)**0.5
        net_ot_approx = m_ot + p[2]*l_ot_approx
        if net_ot_approx/m_ot > 0.1:
            dz = abs(dz)
            # rx = ry = rz = 0

        rx, ry, rz = 1e-7 * p_unit_rot

        u[0:6] = [dx, dy, dz, rx, ry, rz]

        # Compute Floor Plate DOF initial guesses
        for dof_map, plate in zip(self.floor_plate_dofs, self.floor_plates):
            dzp = dz + plate.yc * rx - plate.xc * ry  # z-displacement at (xc,yc)

            # # Handle Vertical Releases
            # if (plate.release_zp and not plate.release_zn and dzp >=0) or (plate.release_zn and not plate.release_zp and dzp <0):
            #     if plate.release_mx:
            #         u[dof_map[3]] = 0
            #     if plate.release_my:
            #         u[dof_map[4]] = 0
            #     return u

            # Handle Independent Vertical DOF (Triggered when both zn and zp releases are present)
            if plate.release_zp and plate.release_zn:
                # dzp = max(0, 0.5*dzp)
                dzp = 0.5 * abs(dzp)
                # dzp = 1e-7
                u[dof_map[2]] = dzp

            dza = 0.5 * abs(dzp)  # Initial anchor point "small tension" target value

            # Handle Rotation Releases
            if plate.release_mx and plate.release_my:
                dx = plate.xy_anchors[:, 0] - plate.x0
                dy = plate.xy_anchors[:, 1] - plate.y0
                delta = dza - dzp

                # Safe division: skip entries with zero denominator
                with np.errstate(divide='ignore', invalid='ignore'):
                    ryp = np.divide(-0.5 * delta, dx, where=dx != 0)
                    rxp = np.divide(0.5 * delta, dy, where=dy != 0)

                # Filter out any NaNs that might remain due to zero division
                ryp = ryp[~np.isnan(ryp)]
                rxp = rxp[~np.isnan(rxp)]

                if rxp.size > 0:
                    u[dof_map[3]] = np.mean(rxp)
                if ryp.size > 0:
                    u[dof_map[4]] = np.mean(ryp)

            elif plate.release_mx and not plate.release_my:
                dy = plate.xy_anchors[:, 1] - plate.y0
                dx = plate.xy_anchors[:, 0] - plate.x0
                delta = dza - dzp + ry * dx

                with np.errstate(divide='ignore', invalid='ignore'):
                    rxp = np.divide(delta, dy, where=dy != 0)

                rxp = rxp[~np.isnan(rxp)]
                if rxp.size > 0:
                    u[dof_map[3]] = np.mean(rxp)

            elif plate.release_my and not plate.release_mx:
                dy = plate.xy_anchors[:, 1] - plate.y0
                dx = plate.xy_anchors[:, 0] - plate.x0
                delta = dza - dzp - rx * dy

                with np.errstate(divide='ignore', invalid='ignore'):
                    ryp = np.divide(delta, -dx, where=dx != 0)

                ryp = ryp[~np.isnan(ryp)]
                if ryp.size > 0:
                    u[dof_map[4]] = np.mean(ryp)
        return u

    def initial_stiffness_matrix_UNUSED(self):
        """ Calculates an initial stiffness matrix based on full compression of bearing areas and tension of anchors."""
        u = np.zeros(self.n_dof)
        k = sum(element.get_element_stiffness_matrix(u, initial=True) for element in self.floor_plates)

        if self.base_straps:
            k_straps = sum(element.get_element_stiffness_matrix(u, initial=True) for element in self.base_straps)
            k += k_straps

        # add brackets to this function and or modify update_stiffness_matrix
        # k_brackets = [element.get_element_stiffness_matrix(u)[0] for element in self.wall_brackets if
        #               element.get_element_stiffness_matrix(u)[0] is not None]
        #
        # if k_brackets:
        #     k[:6, :6] += np.sum(k_brackets, axis=0)

        return k

    def update_stiffness_matrix(self, u):
        k = np.zeros((self.n_dof, self.n_dof))

        k += sum(element.get_element_stiffness_matrix(u) for element in self.floor_plates)
        k += sum(element.get_element_stiffness_matrix(u) for element in self.base_straps)

        k_brackets = [element.get_element_stiffness_matrix(u)[0] for element in self.wall_brackets if
                      element.get_element_stiffness_matrix(u)[0] is not None]

        if k_brackets:
            k[:6, :6] += np.sum(k_brackets, axis=0)

        return k

    def equilibrium_residual(self, u, p, stiffness_update_threshold=0.01, penalty_factor=1e3, verbose=False):
        """Returns the residual for the equilibrium equation p = ku
        stiffness_update_threshold indicates percentage of dof norm change at which stiffness matrix should be updated.
        pentalty factor is applided to zero-force dof residuals to ensure better convergence"""

        load_vector_weights = np.ones(self.n_dof)
        load_vector_weights += 20 * np.array(self.disp_dofs)

        disp_vector_weights = np.ones(self.n_dof)
        # disp_vector_weights[0:3] = [0.1, 0.1, 0.1]

        # Update Stiffness Matrix for large change in norm of u or if any base anchors have reversed signs
        if self.K is None or self._u_previous is None:
            update_K = True  # First iteration, always update K
        else:
            norm_change = np.linalg.norm(disp_vector_weights * u - disp_vector_weights * self._u_previous) / \
                          (np.linalg.norm(disp_vector_weights * self._u_previous) + 1e-12)
            norm_change_trigger = norm_change > stiffness_update_threshold

            # num_base_anchors_in_tension = 0 if self.base_anchors is None else sum(
            #     Utilities.vertical_point_displacements(self.base_anchors.xy_anchors, u) > 0)
            # anchor_change_trigger = num_base_anchors_in_tension != self._num_base_anchors_in_tension

            update_K = norm_change_trigger  # or anchor_change_trigger
            # self._num_base_anchors_in_tension = num_base_anchors_in_tension.copy()

        self.residual_call_count+=1
        if self.residual_call_count % 10 == 0:
            upda_K = True

        if update_K:
            self.K = self.update_stiffness_matrix(u)
            self._u_previous = u.copy()

        # Penalty to prevent "run-away" dofs
        disp_limit = 10
        rotation_limit = 1

        disp_indices = np.where(np.array(self.disp_dofs) == 1)[0]
        rot_indices = np.where(np.array(self.disp_dofs) == 0)[0]

        # Extract corresponding u values
        disp_values = u[disp_indices]
        rot_values = u[rot_indices]

        # Apply penalty logic
        disp_excess = np.maximum(0, np.abs(disp_values) - disp_limit)
        rot_excess = np.maximum(0, np.abs(rot_values) - rotation_limit)

        # Define a quadratic penalty for each DOF
        penalty_disp = 1e6 * disp_excess ** 2
        penalty_rot = 1e6 * rot_excess ** 2
        penalty_vector = np.zeros(self.n_dof)
        penalty_vector[disp_indices] = penalty_disp
        penalty_vector[rot_indices] = penalty_rot

        # Pentalty on zero-force dofs
        zero_force_mask = np.isclose(p, 0, 1e-10)
        load_vector_weights[zero_force_mask] *= penalty_factor
        residual = load_vector_weights * (np.dot(self.K, u) - p) + penalty_vector
        if verbose:
            print(f'Norm: {np.linalg.norm(residual):.3e}, Update K: {update_K}, Penalties: {penalty_vector}')
        return residual

    def solve_equilibrium_Newton_Krylov(self, u_init, p):
        """ Solve the nonlinear equilibrium equation with Newton-Krylov solver. """
        sol = newton_krylov(lambda u: self.equilibrium_residual(u, p), u_init, method='lgmres')
        return sol, np.linalg.norm(self.equilibrium_residual(sol, p)) < 1e-6

    def solve_equilibrium(self, u_init, p):
        methods = ['hybr']
        for method in methods:
            self._u_previous = None
            self.residual_call_count = 0
            res = root(self.equilibrium_residual, u_init, args=p, method=method, #xtol=1e-8,
                       options={'maxfev': 200,'xtol':1e-6})
            if res.success:
                # print(f'Success with {res.nfev} function calls')
                break
        sol = res.x
        return sol, res.success

    def solve_equilibrium_timed(self, u_init, p, timeout=10):
        # Wrap the instance method to make it picklable for multiprocessing
        residual_func = partial(self.equilibrium_residual, verbose=False)

        # Call timeout-enabled solver
        sol, success = solve_equilibrium_with_timeout(
            residual_func,
            u_init,
            p,
            timeout=timeout,
            method='hybr',
            options={'maxfev': 200, 'xtol': 1e-6}
        )
        return sol, success

    def solve_equilibrium_broyden_OLD(self, u_init, p):
        sol = broyden1(lambda u_i: self.equilibrium_residual(u_i, p), u_init, f_tol=1e-1, verbose=True)
        success = np.linalg.norm(self.equilibrium_residual(sol, p)) < 1e-1
        return sol, success

    def analyze_model(self, initial_solution_cache, verbose=False):
        """Applies Horizontal Loads at all angles, solves for equilibrium and stores the solution displacements"""
        # Initialize Analysis Range
        num_theta_z = 4 * 8 + 1
        self.theta_z = np.linspace(0, 2 * math.pi, num_theta_z)
        self.converged = []
        self.equilibrium_solutions = np.zeros((self.n_dof, len(self.theta_z)))

        try_previous_converged = False
        u_prev = np.zeros(self.n_dof)
        # Analysis Attempt with Initial DOF Guesses
        for i, t in enumerate(self.theta_z):
            p = self.get_load_vector(t)

            u_tries = self.get_initial_dof_guess(t)
            if  try_previous_converged:
                u_tries = [u_prev] + u_tries
            if initial_solution_cache is not None:
                u_tries = [initial_solution_cache[:,i]] + u_tries

            success = False
            for j, u_init in enumerate(u_tries):
                # if verbose:
                #     print(f'Theta {np.degrees(t):.2f} trying with u guess {j}')
                sol, success = self.solve_equilibrium(u_init, p)
                if success:
                    if verbose:
                        print(f'Theta {np.degrees(t):.0f} success with initial u guess {j}')
                    self.equilibrium_solutions[:, i] = sol
                    u_prev = sol
                    try_previous_converged = True
                    break

            if not success and verbose:
                print(f'Theta {np.degrees(t):.0f} UNCONVERGED with initial u guess {j}')

            self.converged.append(success)

        # Secondary Analysis Attempt to "Fill-in" Failed Convergence Points
        while any(self.converged) and not all(self.converged):
            if verbose:
                print('Attempting to reanalyze unconverged points...')
            unconverged = [i for i, suc in enumerate(self.converged) if not suc]
            new_converge_found = False
            for idx in unconverged:
                t = self.theta_z[idx]
                p = self.get_load_vector(t)
                u_tries = self.get_interpolated_dof_guess(idx)
                success = False
                for j, u_init in enumerate(u_tries):
                    sol, success = self.solve_equilibrium(u_init, p)
                    if success:
                        if verbose:
                            print(f'Theta {np.degrees(t):.0f} success with interpolated u guess {j}')
                        self.equilibrium_solutions[:, idx] = sol
                        new_converge_found = True
                        break
                self.converged[idx] = success
                if not success and verbose:
                    print(f'Theta {np.degrees(t):.0f} UNCONVERGED with interpolated u guess {j}')
            if not new_converge_found:
                break
        # Compile arrays of element forces
        if not any(self.converged):
            raise Exception(f'Item {self.equipment_id} failed to converge to any solutions. Check geoemtry definition.')
        self.get_element_results()

    def get_element_results(self):
        # Initialize Base Anchor Results Array
        if self.base_anchors is not None:
            self.base_anchors.anchor_forces = np.zeros((len(self.base_anchors.xy_anchors), len(self.theta_z), 3))

        # Initialize Wall Bracket Results Array
        if self.wall_brackets:
            self.wall_bracket_forces = np.zeros(
                (len(self.wall_brackets), len(self.theta_z), 3))  # n_bracket, n_theta, (N,P,Z)-forces

        # Initialize Wall Anchors Results Array
        ''' When concrete/cmu anchors are used, the anchor's object is assigned as an attribute of the model.
                    When sms/wood anchors are used, the anchor's objects are assigned as an attribute of the backing elements.
                    This is because concrete anchors must consider spacing requirements (and thus include all wall anchors
                    in a single object). However, sms anchors must consider pyring conditions caused by unistrut backing,
                    and so must be considered separately backing-by-backing.'''
        for position, wall_anchors in self.wall_anchors.items():
            if wall_anchors is not None:
                wall_anchors.anchor_forces = np.zeros((wall_anchors.xy_anchors.shape[0], len(self.theta_z), 3))

        for backing in self.wall_backing:
            if backing.anchors_obj is not None:
                backing.anchors_obj.anchor_forces = np.zeros((backing.anchors_obj.xy_anchors.shape[0],len(self.theta_z), 3))

        for i, sol in enumerate(self.equilibrium_solutions.T):

            # Update Element Resultants
            self.update_element_resultants(sol)

            # Extract Base Anchor Results
            if self.base_anchors is not None:
                self.base_anchors.anchor_forces[:, i, 0] = np.concatenate(
                    [el.anchor_result['tension'] for el in self.floor_plates if el.n_anchor > 0], axis=0)
                self.base_anchors.anchor_forces[:, i, 1] = np.concatenate(
                    [el.anchor_result['vx'] for el in self.floor_plates if el.n_anchor > 0], axis=0)
                self.base_anchors.anchor_forces[:, i, 2] = np.concatenate(
                    [el.anchor_result['vy'] for el in self.floor_plates if el.n_anchor > 0], axis=0)

            # Extract Wall Bracket Results
            for b, element in enumerate(self.wall_brackets):
                self.wall_bracket_forces[b, i, 0] = element.bracket_forces['fn']
                self.wall_bracket_forces[b, i, 1] = element.bracket_forces['fp']
                self.wall_bracket_forces[b, i, 2] = element.bracket_forces['fz']

            # Extract Wall Anchor Results



            for wall_loc, wall_anchors in self.wall_anchors.items():
                if wall_anchors is not None:
                    forces = np.concatenate(
                        [el.anchor_forces for el in [self.wall_backing[idx] for idx in self.backing_indices[wall_loc]]],
                        axis=0)
                    wall_anchors.anchor_forces[:, i, :] = forces

            for backing in self.wall_backing:
                if backing.anchors_obj is not None:
                    backing.anchors_obj.anchor_forces[:, i, :] = backing.anchor_forces

        self.get_governing_solutions()

    def get_governing_solutions(self):
        """Post Processes the element forces to identify the maximum demand and angle of loading"""

        def get_max_demand(el_vs_theta_matrix):
            return np.unravel_index(np.argmax(el_vs_theta_matrix), el_vs_theta_matrix.shape)

        # Base Anchors
        if self.base_anchors is not None:
            # Tension
            matrix = self.base_anchors.anchor_forces[:, self.converged, 0]
            idx_anchor, idx_theta = get_max_demand(matrix)
            self.governing_solutions['base_anchor_tension']['theta_z'] = self.theta_z[self.converged][idx_theta]
            self.governing_solutions['base_anchor_tension']['sol'] = self.equilibrium_solutions[:,self.converged][:, idx_theta]

            # Shear  # todo: this may need to be modified for the four shear cases xp, xn, yp, yn
            matrix = (self.base_anchors.anchor_forces[:, self.converged, 1] ** 2 + self.base_anchors.anchor_forces[:, self.converged,
                                                                      2] ** 2) ** 0.5
            idx_anchor, idx_theta = get_max_demand(matrix)
            self.governing_solutions['base_anchor_shear']['theta_z'] = self.theta_z[self.converged][idx_theta]
            self.governing_solutions['base_anchor_shear']['sol'] = self.equilibrium_solutions[:,self.converged][:, idx_theta]

        # Wall Brackets
        if self.wall_brackets:
            # Tension
            matrix = self.wall_bracket_forces[:, self.converged, 0]
            idx_anchor, idx_theta = get_max_demand(matrix)
            self.governing_solutions['wall_bracket_tension']['theta_z'] = self.theta_z[self.converged][idx_theta]
            self.governing_solutions['wall_bracket_tension']['sol'] = self.equilibrium_solutions[:,self.converged][:, idx_theta]

            # Shear
            matrix = (self.wall_bracket_forces[:, self.converged, 1] ** 2 + self.wall_bracket_forces[:, self.converged, 2] ** 2) ** 0.5
            idx_anchor, idx_theta = get_max_demand(matrix)
            self.governing_solutions['wall_bracket_shear']['theta_z'] = self.theta_z[self.converged][idx_theta]
            self.governing_solutions['wall_bracket_shear']['sol'] = self.equilibrium_solutions[:,self.converged][:, idx_theta]

            # Wall Anchors
            t_max = -np.inf
            v_max = -np.inf
            all_wall_anchors = [anchors_obj for wall_position, anchors_obj in self.wall_anchors.items()
                                if anchors_obj is not None] + [b.anchors_obj for b in self.wall_backing if b.anchors_obj is not None]
            for wall_anchors in all_wall_anchors:
                # Tension
                matrix = wall_anchors.anchor_forces[:, self.converged, 0]
                idx_anchor, idx_theta = get_max_demand(matrix)
                t_new = matrix[idx_anchor, idx_theta]
                if t_new > t_max:
                    t_max = t_new.copy()
                    self.governing_solutions['wall_anchor_tension']['theta_z'] = self.theta_z[self.converged][idx_theta]
                    self.governing_solutions['wall_anchor_tension']['sol'] = self.equilibrium_solutions[:,self.converged][:, idx_theta]

                # Shear
                matrix = (wall_anchors.anchor_forces[:, self.converged, 1] ** 2 + wall_anchors.anchor_forces[:, self.converged,
                                                                     2] ** 2) ** 0.5
                idx_anchor, idx_theta = get_max_demand(matrix)
                v_new = matrix[idx_anchor, idx_theta]
                if v_new > v_max:
                    v_max = v_new.copy()
                self.governing_solutions['wall_anchor_shear']['theta_z'] = self.theta_z[self.converged][idx_theta]
                self.governing_solutions['wall_anchor_shear']['sol'] = self.equilibrium_solutions[:,self.converged][:, idx_theta]

    def update_element_resultants(self, sol):
        for element in self.floor_plates:
            # todo: [Refinement] revise the three methods below, so that matrices are not re-computed for each method. Consider adding an update_state(u) method to be called before the get_resultant methods
            element.get_connection_forces(sol)
            element.get_anchor_resultants(sol)
            element.get_compression_resultants(sol)

            if element.connection:

                element.connection.get_anchor_forces(*Utilities.transform_forces(
                    element.connection_forces, element.connection.normal_vector))
                element.connection.anchors_obj.check_anchors(self.asd_lrfd_ratio)

        for element in self.base_straps:
            element.check_brace(sol, self.asd_lrfd_ratio)

        for element in self.wall_brackets:
            element.get_element_forces(sol)
            element.check_brackets()

            if element.connection:
                element.compute_connection_forces()
                element.connection.get_anchor_forces(*element.connection_forces)
                element.connection.anchors_obj.check_anchors(self.asd_lrfd_ratio)

        for wall_loc, indices in self.backing_indices.items():
            for el in [self.wall_backing[idx] for idx in indices]:
                el.get_anchor_forces([self.wall_brackets[i] for i in el.bracket_indices])


class FloorPlateElement:
    def __init__(self, bearing_boundaries, xy_anchors=None,
                 x0=0.0, y0=0.0, z0=0.0,
                 xc=0.0, yc=0.0, zc=0.0,
                 release_mx=False, release_my=False, release_mz=False,
                 release_xp=False, release_xn=False,
                 release_yp=False, release_yn=False, release_zp=False, release_zn=False,
                 E_base=None, poisson=None):

        # Geometry
        self.bearing_boundaries = bearing_boundaries

        if xy_anchors is None:
            self.xy_anchors = np.array([])
            self.n_anchor = 0
        else:
            self.xy_anchors = xy_anchors
            self.n_anchor = np.shape(self.xy_anchors)[0]

        # Material Properties
        self.anchor_shear_stiffness = None
        self.anchor_tension_stiffness = None
        self.E_base = E_base  # Elastic modulus of base material
        self.poisson = poisson  # Poisson ratio of base material

        # Attachment Fasteners
        self.connection = None

        # DOF and Constraint Parameters
        self.x0 = x0  # Inflection point of element
        self.y0 = y0  # Inflection point of element
        self.z0 = z0  # Inflection point of element

        self.xc = xc  # Connection Point of element
        self.yc = yc  # Connection Point of element
        self.zc = zc  # Connection Point of element

        self.release_mx = release_mx  # Moment release at attachment point about x-axis
        self.release_my = release_my  # Moment release at attachment point about y-axis
        self.release_mz = release_mz
        self.release_xp = release_xp
        self.release_xn = release_xn
        self.release_yp = release_yp
        self.release_yn = release_yn
        self.release_zp = release_zp
        self.release_zn = release_zn

        # Stiffness Matrix Parameters
        self.C = None  # Constraint Matrix (global DOFs to local DOFs)
        self.B = np.array([[1, 0, 0, 0, 0, 0],  # Static Conversion Matrix (dofs to connection point)
                           [0, 1, 0, 0, 0, 0],
                           [0, 0, 1, 0, 0, 0],
                           [0, (self.zc - self.z0), -(self.yc - self.y0), 1, 0, 0],
                           [-(self.zc - self.z0), 0, (self.xc - self.x0), 0, 1, 0],
                           [(self.yc - self.y0), -(self.xc - self.x0), 0, 0, 0, 1]])

        # self.ka_template = None  # ka with "kz" value set to 1
        # self.ka = None  # Stiffness sub-matrix for base_anchors
        # self.kb = None  # Stiffness sub-matrix for bearing areas

        # State Parameters
        self.cz_result = {}  # Dictionary of compression zone geometries and resultants
        self.anchor_result = {}  # Dictionary of anchor force resultants
        self.nodal_forces = None
        self.connection_forces = None

    def set_dof_constraints(self, n_dof, dof_map):
        """Initializes static (displacement-independent) matrices given the total global dofs and dof_map.
         dof_map is a three element array indicating the index of the global DOFs at each local dofs"""

        """ Returns the constraint matrix (6 x n_dof) relating global DOFs to 6 element DOFs"""

        ''' Presence of vertical releases in both z+ and z- directions results in uncoupling of base plate vertical
        translation as a unique degree of freedom.
        Shear releases do not result in unique degrees of freedom, but rather are handled by modifying the shear
        of the anchors so that resulting plots will show floor plates kinematically constrained to the unit without
        imparting stiffness.'''

        self.C = np.zeros((6, n_dof))
        self.C[0, 0:6] = [1, 0, 0, 0, self.z0, -self.y0]
        self.C[1, 0:6] = [0, 1, 0, -self.z0, 0, self.x0]
        if dof_map[2] != 2:
            self.C[2, dof_map[2]] = 1
        else:
            self.C[2, 0:6] = [0, 0, 1, self.yc, -self.xc, 0]
        # todo: verify and revise. There should be coupling for rotation dofs when vertical dof is free on baseplate
        self.C[3, dof_map[3]] = 1
        self.C[4, dof_map[4]] = 1
        self.C[5, dof_map[5]] = 1

    def set_anchor_properties(self, anchors_object):
        """ Updates the anchor stiffness array [kx, ky, kz], and base material properties"""
        self.anchor_shear_stiffness = anchors_object.Kv
        self.anchor_tension_stiffness = anchors_object.K

    def set_sms_attachment_UNUSED(self, xy_anchors, ):
        self.connection = SMSAnchors(xy_anchors=xy_anchors)
        self.connection.set_sms_properties()  # todo [Attachments, pass actual equipment values here]

    def update_anchor_stiffness_matrix(self, u, initial=False):
        """ Computes and returns the anchor stiffness matrix without modifying self. """

        # Initialize stiffness values
        kx = np.full(self.n_anchor, self.anchor_shear_stiffness)
        ky = np.full(self.n_anchor, self.anchor_shear_stiffness)
        kz = np.full(self.n_anchor, self.anchor_tension_stiffness)

        if not initial:
            # Compute displacements
            xyz_anchors = np.column_stack((self.xy_anchors, np.zeros(self.n_anchor)))
            delta = Utilities.compute_point_displacements(xyz_anchors, self.C @ u, x0=self.x0, y0=self.y0, z0=self.z0)

            # Vectorized release conditions
            kx *= ~((self.release_xp & (delta[:, 0] >= 0)) | (self.release_xn & (delta[:, 0] < 0)))
            ky *= ~((self.release_yp & (delta[:, 1] >= 0)) | (self.release_yn & (delta[:, 1] < 0)))
            if not(self.release_zn and self.release_zp):  # both releasees triggers independent vertical plate dof
                kz *= (not self.release_zp) & (delta[:, 2] >= 0)
            else:
                kz *= (delta[:, 2] >= 0)

        # Vectorized matrix assembly
        x, y = self.xy_anchors[:, 0], self.xy_anchors[:, 1]
        ka = np.zeros((6, 6, self.n_anchor))  # Local variable instead of modifying self.ka

        ka[0, 0, :] = kx
        ka[0, 4, :] = ka[4, 0, :] = -kx * self.z0
        ka[0, 5, :] = ka[5, 0, :] = -kx * (y - self.y0)

        ka[1, 1, :] = ky
        ka[1, 3, :] = ka[3, 1, :] = ky * self.z0
        ka[1, 5, :] = ka[5, 1, :] = ky * (x - self.x0)

        ka[2, 2, :] = kz
        ka[2, 3, :] = ka[3, 2, :] = kz * (y - self.y0)
        ka[2, 4, :] = ka[4, 2, :] = -kz * (x - self.x0)

        ka[3, 3, :] = kz * (y - self.y0) ** 2 + ky * self.z0 ** 2
        ka[3, 4, :] = ka[4, 3, :] = -kz * (x - self.x0) * (y - self.y0)
        ka[3, 5, :] = ka[5, 3, :] = ky * (x - self.x0) * self.z0

        ka[4, 4, :] = kz * (x - self.x0) ** 2 + kx * self.z0 ** 2
        ka[4, 5, :] = ka[5, 4, :] = kx * (y - self.y0) * self.z0

        ka[5, 5, :] = ky * (x - self.x0) ** 2 + kx * (y - self.y0) ** 2

        return ka

    def get_compression_zone_properties(self, u, full_bearing=False):
        """Returns the geometric properties of the compression zones for given dof values, u"""
        # Identify area of bearing elements that are in compression
        # Loop trough each bearing boundary, then loop through each compression area and assemble list.

        # Case 1, no moment releases present:
            # Turn off all brearing
        # Case 2, moment releases are present:
            # Turn off bearing only if node is negatively displaced

        if (self.release_zn and not self.release_zp) and not (self.release_mx or self.release_my):
            release_bearing = True
        elif (self.release_zn and not self.release_zp) and (Utilities.vertical_point_displacements(np.array([[self.x0,self.y0,self.z0]]), u)[0]<=0):
            release_bearing = True
        else:
            release_bearing = False

        if release_bearing:
            compression_boundaries = []
        elif full_bearing:
            u0 = np.array([0, 0, -1e-8, 0, 0, 0])
            compression_boundaries = [
                compression_boundary
                for boundary in self.bearing_boundaries
                for compression_boundary in
                Utilities.bearing_area_in_compression(boundary, u0, x0=self.x0, y0=self.y0)]
        else:
            compression_boundaries = [
                compression_boundary
                for boundary in self.bearing_boundaries
                for compression_boundary in
                Utilities.bearing_area_in_compression(boundary, self.C @ u, x0=self.x0, y0=self.y0)]

        # Get stiffness and inertial properties for compression areas
        n_boundaries = len(compression_boundaries)
        areas = np.empty(n_boundaries)
        centroids = np.empty((n_boundaries, 2))
        Ixx_list = np.empty(n_boundaries)
        Iyy_list = np.empty(n_boundaries)
        Ixy_list = np.empty(n_boundaries)
        beta_list = np.empty(n_boundaries)

        for i, vertices in enumerate(compression_boundaries):
            area, centroid, Ixx, Iyy, Ixy = Utilities.polygon_properties(vertices)
            beta = Utilities.effective_indenter_stiffness(area, self.E_base, self.poisson)

            areas[i] = area
            centroids[i] = centroid
            Ixx_list[i] = Ixx
            Iyy_list[i] = Iyy
            Ixy_list[i] = Ixy
            beta_list[
                i] = beta if not full_bearing else beta / 10  # heuristic to reduce stiffness for initial stiffness matrix

        self.cz_result = {'compression_boundaries': compression_boundaries,
                          'areas': areas,
                          'centroids': centroids,
                          'Ixx': Ixx_list,
                          'Iyy': Iyy_list,
                          'Ixy': Ixy_list,
                          'beta': beta_list}

    def update_bearing_stiffness_matrix(self, u, initial=False):
        """ Computes and returns the bearing stiffness matrix without modifying self. """

        # Get compression zones (consider caching this if u doesn't change significantly)
        if initial:
            self.get_compression_zone_properties(None, full_bearing=True)
        else:
            self.get_compression_zone_properties(u)

        # If there are no compression zones, return a zero matrix
        if len(self.cz_result['areas']) == 0:
            return np.zeros((6, 6))

        # Precompute compression zone stiffness matrices
        kb_list = [
            self.compression_zone_matrix(A, x_bar, y_bar, Ixx, Iyy, Ixy, beta, self.x0, self.y0)
            for A, (x_bar, y_bar), Ixx, Iyy, Ixy, beta in zip(
                self.cz_result['areas'],
                self.cz_result['centroids'],
                self.cz_result['Ixx'],
                self.cz_result['Iyy'],
                self.cz_result['Ixy'],
                self.cz_result['beta'])
        ]

        # Efficiently sum all stiffness matrices
        kb = np.sum(kb_list, axis=0)

        return kb

    @staticmethod
    def compression_zone_matrix(A, x_bar, y_bar, Ixx, Iyy, Ixy, beta, x0=0.0, y0=0.0):
        """Returns the local bearing stiffness matrix for a single compression zone"""
        k_cz = np.zeros((6, 6))

        # Assign values according to the LaTeX expressions (note the zero-indexing in Python)
        k_cz[2, 2] = beta * A
        k_cz[2, 3] = k_cz[3, 2] = beta * (y_bar - y0) * A
        k_cz[2, 4] = k_cz[4, 2] = -beta * (x_bar - x0) * A
        k_cz[3, 3] = beta * (Ixx - 2 * y0 * y_bar * A + y0 ** 2 * A)
        k_cz[3, 4] = k_cz[4, 3] = -beta * (Ixy - y0 * x_bar * A - x0 * y_bar * A + x0 * y0 * A)
        k_cz[4, 4] = beta * (Iyy - 2 * x0 * x_bar * A + x0 ** 2 * A)

        return k_cz

    def get_element_stiffness_matrix_OLD(self, u):
        """ Assembles the element stiffness matrix from anchor and bearing matrices,
        Here, this refers to the stiffness matrix relating global dofs to global forces"""
        has_anchors = len(self.xy_anchors) > 0
        has_bearing = len(self.bearing_boundaries) > 0

        if len(self.xy_anchors) > 0 and len(self.bearing_boundaries) > 0:
            self.update_anchor_stiffness_matrix(u)
            self.update_bearing_stiffness_matrix(u)
            k_element = self.C.T @ (np.sum(self.ka, axis=2) + self.kb) @ self.C
        elif len(self.bearing_boundaries) > 0:
            self.update_bearing_stiffness_matrix(u)
            k_element = self.C.T @ self.kb @ self.C
        elif len(self.xy_anchors) > 0:
            self.update_anchor_stiffness_matrix(u)
            k_element = self.C.T @ np.sum(self.ka, axis=2) @ self.C
        else:
            ndof = self.C.shape[1]
            k_element = np.zeros((ndof, ndof))

        return k_element

    def get_element_stiffness_matrix(self, u, initial=False):
        """ Efficiently assembles the element stiffness matrix without modifying self. """
        has_anchors = len(self.xy_anchors) > 0
        ka = self.update_anchor_stiffness_matrix(u, initial=initial) if has_anchors else np.zeros((6, 6, self.n_anchor))
        kb = self.update_bearing_stiffness_matrix(u, initial=initial)

        # Precompute anchor stiffness sum
        k_anchor = np.sum(ka, axis=2) if has_anchors else np.zeros((6, 6))

        # Use matrix multiplication only once
        k_element = self.C.T @ (k_anchor + kb) @ self.C

        return k_element

    def get_compression_resultants(self, u):
        """Computes the compression zone resultant forces and centroids"""
        self.update_bearing_stiffness_matrix(u)

        if len(self.cz_result['compression_boundaries']) != 0:

            x0 = self.x0
            y0 = self.y0

            fz = np.zeros(len(self.cz_result['compression_boundaries']))
            mx = np.zeros(len(self.cz_result['compression_boundaries']))
            my = np.zeros(len(self.cz_result['compression_boundaries']))

            # get compression zones
            for i, (A, (x_bar, y_bar), Ixx, Iyy, Ixy, beta) in enumerate(zip(
                    self.cz_result['areas'],
                    self.cz_result['centroids'],
                    self.cz_result['Ixx'],
                    self.cz_result['Iyy'],
                    self.cz_result['Ixy'],
                    self.cz_result['beta'])):
                kb_cz = self.compression_zone_matrix(A, x_bar, y_bar, Ixx, Iyy, Ixy, beta, x0, y0)

                p = kb_cz @ self.C @ u

                fz[i] = p[2]
                mx[i] = p[3]
                my[i] = p[4]

            self.cz_result['fz'] = fz
            self.cz_result['resultant_centroids'] = np.column_stack((-my / fz + x0, mx / fz + y0))
        else:
            self.cz_result['fz'] = []
            self.cz_result['resultant_centroids'] = []

    def get_anchor_resultants(self, u):
        """Computes anchor forces"""
        if self.n_anchor > 0:
            ka = self.update_anchor_stiffness_matrix(u)
            p = np.einsum('ijk,jl,l->ik', ka, self.C, u)
            vx = p[0, :]
            vy = p[1, :]
            t = p[2, :]
            tension_resultant = t.sum()
            mx = p[3, :]
            my = p[4, :]

            if not np.isclose(t.sum(), 0):
                resultant_centroid = np.array([-my.sum() / t.sum(), mx.sum() / t.sum()])
            else:
                resultant_centroid = np.NaN
        else:
            vx = None
            vy = None
            t = None
            tension_resultant = None
            resultant_centroid = None

        self.anchor_result = {
            'vx': vx,
            'vy': vy,
            'tension': t,
            'tension_resultant': tension_resultant,
            'resultant_centroid': resultant_centroid}

    def get_nodal_forces_OLD(self, u):
        """ Returns the forces at the attachment point of the floor plate element"""
        # Assemble element stiffness matrix (relating global dofs to element basic forces
        if len(self.xy_anchors) > 0 and len(self.bearing_boundaries) > 0:
            self.update_anchor_stiffness_matrix(u)
            self.update_bearing_stiffness_matrix(u)
            k_element = (np.sum(self.ka, axis=2) + self.kb) @ self.C
        elif len(self.bearing_boundaries) > 0:
            self.update_bearing_stiffness_matrix(u)
            k_element = self.kb @ self.C
        elif len(self.xy_anchors) > 0:
            self.update_anchor_stiffness_matrix(u)
            k_element = np.sum(self.ka, axis=2) @ self.C
        else:
            ndof = self.C.shape[1]
            k_element = np.zeros((6, ndof))

        self.nodal_forces = k_element @ u

    def get_nodal_forces(self, u):
        """ Computes and returns the forces at the attachment point of the floor plate element """

        # Check if anchors and/or bearings exist
        has_anchors = len(self.xy_anchors) > 0
        has_bearings = len(self.bearing_boundaries) > 0

        # Compute anchor and bearing stiffness matrices
        ka = self.update_anchor_stiffness_matrix(u) if has_anchors else np.zeros((6, 6, self.n_anchor))
        kb = self.update_bearing_stiffness_matrix(u) if has_bearings else np.zeros((6, 6))

        # Precompute anchor stiffness sum
        k_anchor = np.sum(ka, axis=2) if has_anchors else np.zeros((6, 6))

        # Compute element stiffness matrix
        if has_anchors and has_bearings:
            k_element = (k_anchor + kb) @ self.C
        elif has_bearings:
            k_element = kb @ self.C
        elif has_anchors:
            k_element = k_anchor @ self.C
        else:
            ndof = self.C.shape[1]
            k_element = np.zeros((6, ndof))

        # Compute nodal forces
        self.nodal_forces = k_element @ u

        return

    def get_connection_forces(self, u):
        self.get_nodal_forces(u)

        # Sum nodal forces at connection point
        self.connection_forces = self.B @ self.nodal_forces

    def check_prying_thickness(self):
        """ update element resultants should be called before calling this method"""
        if self.release_mx == 'Check Prying':
            Mu = self.nodal_forces[3]
        elif self.release_my == 'Check Prying':
            Mu = self.nodal_forces[4]
        else:
            return False

        phi = 0.9
        self.tnp = (4*Mu/(phi*self.yield_width*self.fu))
        self.t


if __name__ == '__main__':
    pass


class WallBracketElement:
    def __init__(self, supporting_wall, xyz_0, xyz_f, normal_unit_vector, wall_flexibility, brace_flexibility=None,
                 horizontal_stiffness=None, vertical_stiffness=None, releases=None, xyz_offset=None, e_cxn=0):

        self.xyz_0 = xyz_0  # Coordinate along bracket centerline at connection offset.
        self.xyz_f = xyz_f  # Coordinate at attachment to wall (or wall backing element)
        if xyz_offset is None:
            self.xyz_offset = xyz_0  # Coordinate at attachment to equipment unit
        else:
            self.xyz_offset = xyz_offset

        self.normal = normal_unit_vector
        self.supporting_wall = supporting_wall

        self.length = math.sqrt(sum((f - i) ** 2 for f, i in zip(xyz_offset, xyz_f)))

        # Bracket Properties
        self.bracket_id = None
        self.f_wall = wall_flexibility
        self.f_brace = brace_flexibility
        self.kp = horizontal_stiffness
        self.kz = vertical_stiffness
        self.capacity_to_equipment = None
        self.bracket_capacity = None
        self.capacity_to_backing = None
        self.connection = None
        self.e_cxn = e_cxn  # Eccentricity of bracket center-line normal to connection faying surface
        self.connection_forces = None

        self.releases = releases

        x, y, z = self.xyz_0
        self.C = np.array([[1, 0, 0, 0, z, -y],
                           [0, 1, 0, -z, 0, x],
                           [0, 0, 1, y, -x, 0]])

        nx, ny, nz = self.normal
        self.G = np.array([[nx, ny, 0],
                           [-ny, nx, 0],
                           [0, 0, 1]])

        # self.k_br = None
        # self.K = None

        if brace_flexibility is not None:
            self.set_bracket_properties(brace_flexibility,
                                        horizontal_stiffness, vertical_stiffness)

        self.bracket_forces = {}
        self.tension_dcr = None

    def set_dof_constraints(self, n_dof):
        pass

    def set_bracket_properties(self, bracket_data):
        self.bracket_id = bracket_data['bracket_id']
        self.f_brace = bracket_data['bracket_flexibility']
        self.kp = bracket_data['kp']
        self.kz = bracket_data['kz']
        self.capacity_to_equipment = bracket_data['capacity_to_equipment']
        self.bracket_capacity = bracket_data['bracket_capacity']
        self.capacity_to_backing = bracket_data['capacity_to_backing']

    def get_element_stiffness_matrix_OLD(self, u=None):
        """ Returns a 6x6 stiffness matrix for the primary degrees of freedom"""
        kn = 1 / (self.f_wall + self.f_brace)
        kp = self.kp
        kz = self.kz

        if any(self.releases) and (u is not None):
            delta = np.linalg.multi_dot((self.G, self.C, u))

            # Check Normal Direction
            if delta[0] > 0 and self.releases[0]:
                kn = 0
            if delta[0] < 0 and self.releases[1]:
                kn = 0

            # Check Parallel to Wall axis
            if delta[1] > 0 and self.releases[2]:
                kp = 0
            if delta[1] < 0 and self.releases[3]:
                kp = 0

            # Check Z axis
            if delta[2] > 0 and self.releases[4]:
                kz = 0
            if delta[2] < 0 and self.releases[5]:
                kz = 0

        self.k_br = np.array([[kn, 0, 0],
                              [0, kp, 0],
                              [0, 0, kz]])
        self.K = np.linalg.multi_dot((self.C.T, self.G.T, self.k_br, self.G, self.C))
        return self.K

    def get_element_stiffness_matrix(self, u=None):
        """ Returns a 6x6 stiffness matrix for the primary degrees of freedom without modifying self. """

        # Compute initial stiffness values
        kn = 1 / (self.f_wall + self.f_brace)
        kp = self.kp
        kz = self.kz

        # If releases are defined and u is provided, check for displacement conditions
        if any(self.releases) and (u is not None):
            delta = self.G @ self.C @ u[0:6]  # Efficient matrix multiplication

            kn *= not ((delta[0] > 0 and self.releases[0]) or (delta[0] < 0 and self.releases[1]))
            kp *= not ((delta[1] > 0 and self.releases[2]) or (delta[1] < 0 and self.releases[3]))
            kz *= not ((delta[2] > 0 and self.releases[4]) or (delta[2] < 0 and self.releases[5]))

            # # Vectorized condition checking
            # release_conditions = np.array([
            #     delta[0] > 0 and self.releases[0],  # Normal direction (tension)
            #     delta[0] < 0 and self.releases[1],  # Normal direction (compression)
            #     delta[1] > 0 and self.releases[2],  # Parallel to wall (positive)
            #     delta[1] < 0 and self.releases[3],  # Parallel to wall (negative)
            #     delta[2] > 0 and self.releases[4],  # Z-axis (positive)
            #     delta[2] < 0 and self.releases[5],  # Z-axis (negative)
            # ])
            #
            # stiffness_values = np.array([kn, kn, kp, kp, kz, kz])
            # stiffness_values[release_conditions] = 0  # Zero out affected stiffnesses
            #
            # kn, kn, kp, kp, kz, kz = stiffness_values

        # Construct local stiffness matrix
        k_br = np.diag([kn, kp, kz])

        # Compute global stiffness matrix without modifying self
        K = self.C.T @ self.G.T @ k_br @ self.G @ self.C
        return K, k_br

    def get_element_forces(self, u):
        """Given the"""
        u = u[0:6]
        _, k_br = self.get_element_stiffness_matrix(u=u)
        fn, fp, fz = np.linalg.multi_dot((k_br, self.G, self.C, u))
        self.bracket_forces = {'fn': fn,  # Normal to wall
                               'fp': fp,  # Parallel to Wall (Horizontal)
                               'fz': fz}  # Vertical

    def compute_connection_forces(self):
        """"""
        if not self.connection:
            return

        # Define bracket local axis vectors
        z_vec = np.array([0, 0, 1])
        bracket_n = np.array(self.normal)
        bracket_p = np.cross(z_vec, bracket_n)

        # Define connection local axis vectors
        cxn_norm = np.array(self.connection.normal_vector)
        cxn_x = -bracket_n
        cxn_y = np.cross(cxn_norm, cxn_x)

        # Get Bracket Force Vector in Bracket NPZ coordinates
        f_npz = np.array([self.bracket_forces['fn'],
                          self.bracket_forces['fp'],
                          self.bracket_forces['fz']])

        # Convert Bracket Force Vector to Global XYZ coordinates
        f_xyz = -np.dot(self.G.T, f_npz)  # todo:this line is wrong. verify correct vector conversion

        # Convert Bracket Force Vector to demands on connection in local xyz coordinates
        Vx = np.dot(f_xyz, cxn_x)
        Vy = np.dot(f_xyz, cxn_y)
        N = np.dot(f_xyz, cxn_norm)
        Mx = -Vy * self.e_cxn
        My = Vx * self.e_cxn - N * self.length
        T = Vy * self.length

        self.connection_forces = Vx, Vy, N, Mx, My, T

    def check_brackets(self):
        capacities = [self.capacity_to_equipment, self.bracket_capacity, self.capacity_to_backing]
        governing_capacity = min([item for item in capacities if isinstance(item, (int, float))], default=None)
        self.tension_dcr = self.bracket_forces['fn'] / governing_capacity if governing_capacity else 'OK by inspection'


class RigidHardwareElement:
    """ Represents a rigid, rectangular hardware element with fasteners"""

    def __init__(self, w, h, pz_anchors, xyz_centroid=(0, 0, 0), normal_vector=(1, 0, 0), x_offset=0, y_offset=0):
        self.w = w
        self.h = h
        self.pz_anchors = pz_anchors  # Anchor coordinates in parallel-vertical axes.
        self.n_anchors = pz_anchors.shape[0]
        self.centroid = xyz_centroid
        self.normal_vector = normal_vector
        self.x_offset = x_offset
        self.y_offset = y_offset

        self.Ixx = None
        self.Iyy = None
        self.Ixy = None
        self.Ip = None

        self.N = None
        self.Vx = None
        self.Vy = None
        self.Mx = None
        self.My = None
        self.T = None

        self.get_anchor_group_properties()

    def get_anchor_group_properties(self):
        x_bar, y_bar = np.mean(self.pz_anchors, axis=0)
        dx = self.pz_anchors[:, 0] - x_bar
        dy = self.pz_anchors[:, 1] - y_bar
        self.Ixx = sum(dy ** 2)
        self.Iyy = sum(dx ** 2)
        self.Ixy = sum(dx * dy)
        self.Ip = self.Ixy + self.Iyy

    def set_centroid_forces(self, Vx, Vy, N, Mx, My, T):
        self.N, self.Vx, self.Vy, self.Mx, self.My, self.T = N, Vx, Vy, Mx, My, T

    def get_anchor_forces(self):
        N, Vx, Vy, Mx, My, T = self.N, self.Vx, self.Vy, self.Mx, self.My, self.T
        Ixx = self.Ixx
        Iyy = self.Iyy
        Ixy = self.Ixy
        x = self.pz_anchors[:, 0]
        y = self.pz_anchors[:, 1]

        normal_term = N / self.n_anchors

        if Ixx == 0 and Ixy == 0:
            mx_term = np.zeros(y.shape)
        elif Ixy == 0:
            mx_term = y * (Mx / Ixx)
        else:
            mx_term = y * (Mx * Iyy - My * Ixy) / (Ixx * Iyy - Ixy ** 2)

        if Iyy == 0 and Ixy == 0:
            my_term = np.zeros(x.shape)
        elif Ixy == 0:
            my_term = -x * (My / Iyy)
        else:
            my_term = -x * (My ** Ixx - Mx * Ixy) / (Ixx * Iyy - Ixy ** 2)

        n = normal_term + mx_term + my_term

        Ip = Ixx + Iyy
        if Ip == 0:
            vp = np.ones(self.n_anchors) * Vx / self.n_anchors
            vz = np.ones(self.n_anchors) * Vy / self.n_anchors
        else:
            vp = np.ones(self.n_anchors) * Vx / self.n_anchors - T*y/Ip
            vz = np.ones(self.n_anchors) * Vy / self.n_anchors + T*x/Ip

        return np.column_stack((n, vp, vz))


class WallBackingElement(RigidHardwareElement):
    def __init__(self, w, h, d, pz_anchors, pz_brackets, bracket_indices, supporting_wall,
                 backing_type='Flat', centroid=(0, 0, 0), normal_vector=(1, 0, 0)):
        super().__init__(w, h, pz_anchors, xyz_centroid=centroid, normal_vector=normal_vector)

        self.d = d

        self.pz_brackets = pz_brackets
        self.bracket_indices = bracket_indices
        self.supporting_wall = supporting_wall
        self.backing_type = backing_type

        self.anchor_forces = None
        self.bracket_forces = None

        self.anchors_obj = None

    def get_centroid_forces(self, bracket_list):
        bracket_forces = np.zeros((len(bracket_list), 3))
        for i, bracket in enumerate(bracket_list):
            bracket_forces[i, :] = (bracket.bracket_forces['fn'],
                                    bracket.bracket_forces['fp'],
                                    bracket.bracket_forces['fz'])
        self.bracket_forces = bracket_forces
        # Correct bracket coordinates to place bracket centroid at origin. Calcs assume anchor centroid also at origin
        x_bar, y_bar = np.mean(self.pz_brackets, axis=0)
        N = np.sum(bracket_forces[:, 0])
        Mx = np.sum(bracket_forces[:, 0] * (self.pz_brackets[:, 1] - y_bar))
        My = np.sum(-bracket_forces[:, 0] * (self.pz_brackets[:, 0] - x_bar))
        Vx = bracket_forces[:, 1].sum()
        Vy = bracket_forces[:, 2].sum()
        T = 0  # todo: [Calc Refinement] provide correct torsion from bracket forces here
        self.set_centroid_forces(Vx, Vy, N, Mx, My, T)

    def get_anchor_forces(self, bracket_list):
        self.get_centroid_forces(bracket_list)
        self.anchor_forces = super().get_anchor_forces()


class SMSHardwareAttachment(RigidHardwareElement):
    def __init__(self, w, h, pz_anchors, df_sms, centroid=(0, 0, 0), normal_vector=(1, 0, 0), x_offset=0, y_offset=0):
        super().__init__(w, h, pz_anchors, xyz_centroid=centroid, normal_vector=normal_vector, x_offset=x_offset,
                         y_offset=y_offset)
        self.anchors_obj = SMSAnchors(xy_anchors=pz_anchors, df_sms=df_sms)

    def get_anchor_forces(self, Vx, Vy, N, Mx, My, T):
        self.set_centroid_forces(Vx, Vy, N, Mx, My,
                                 T)  # todo [Connection Fasteners]: Need to fix this to translate global forces to local connection axes
        self.anchors_obj.anchor_forces = super().get_anchor_forces()


class SMSAnchors:
    def __init__(self, wall_data=None, xy_anchors=None, backing_type=None, df_sms=None, condition_x = 'Condition 1',
                 condition_y = 'Condition 1'):
        # SMS capacity table
        self.df_sms = df_sms
        # Geometry
        self.xy_anchors = xy_anchors

        # Sheet Metal Properties
        self.fy = None
        self.gauge = None

        # Screw Properties
        self.screw_size = None

        # Attachment (Capacity) Condition
        self.condition_x = condition_x  # Steel to steel, One layer gyp, Two layer gyp, Prying
        self.condition_y = condition_y

        # Anchor Demands
        self.anchor_forces = None
        self.Tu_max = None
        self.DCR = None
        self.results = {}

        if wall_data is not None:
            self.set_sms_properties(gauge=wall_data['stud_gauge'], fy=wall_data['stud_fy'],
                                    num_gyp=wall_data['num_gyp'], backing_type=backing_type)

        self.conditions = {'Condition 1': {'Label': 'Steel-to-Steel Connection',
                                           'Table': 'Table 1'},
                           'Condition 2': {'Label': 'Single-Layer Gyp. Board (Non-Prying)',
                                           'Table': 'Table 2'},
                           'Condition 3': {'Label': 'Two-layer Gyp. Board (Non-Prying)',
                                           'Table': 'Table 3'},
                           'Condition 4': {'Label': 'Prying Condition',
                                           'Table': 'Table 4'}}

    def set_sms_properties(self, gauge=18, fy=33, num_gyp=0, backing_type=None):
        self.gauge = gauge
        self.fy = fy

    def set_screw_size(self, screw_size):
        self.screw_size = screw_size

    def check_anchors(self, asd_lrfd_ratio):
        # Convert Anchor Forces to ASD
        anchor_forces = self.anchor_forces * asd_lrfd_ratio

        # Determine if anchor_forces is (n x t x 3) or (n x 3)
        if anchor_forces.ndim == 3:
            # Get Maximum Forces for (n x t x 3) case
            tension_demand = anchor_forces[:, :, 0].max()
            shear_demand = ((anchor_forces[:, :, 1] ** 2 + anchor_forces[:, :, 2] ** 2) ** 0.5).max()
            shear_x_demand = np.abs(anchor_forces[:,:,1]).max()
            shear_y_demand = np.abs(anchor_forces[:,:,2]).max()
        elif anchor_forces.ndim == 2:
            # Get Maximum Forces for (n x 3) case
            tension_demand = anchor_forces[:, 0].max()
            shear_demand = ((anchor_forces[:, 1] ** 2 + anchor_forces[:, 2] ** 2) ** 0.5).max()
            shear_x_demand = np.abs(anchor_forces[:, 1]).max()
            shear_y_demand = np.abs(anchor_forces[:, 2]).max()
        else:
            raise ValueError("anchor_forces must be either an (n x t x 3) or (n x 3) array.")

        # Store the results
        self.results['Tension Demand'] = tension_demand
        self.results['Shear Demand'] = shear_demand
        self.results['Shear X Demand'] = shear_x_demand
        self.results['Shear Y Demand'] = shear_y_demand
        self.Tu_max = tension_demand

        filtered_dfx = self.df_sms[
            (self.df_sms['sms_size'] == self.screw_size) &
            (self.df_sms['condition'] == self.condition_x) &
            (self.df_sms['fy'] == self.fy) &
            (self.df_sms['gauge'] == self.gauge)
            ]

        filtered_dfy = self.df_sms[
            (self.df_sms['sms_size'] == self.screw_size) &
            (self.df_sms['condition'] == self.condition_y) &
            (self.df_sms['fy'] == self.fy) &
            (self.df_sms['gauge'] == self.gauge)]

        if not filtered_dfx.empty:
            if np.isnan(filtered_dfx['shear'].values[0]) or np.isnan(filtered_dfx['tension'].values[0]) or np.isnan(filtered_dfy['shear'].values[0]):
                self.results['Shear X Capacity'] = "NA"
                self.results['Shear Y Capacity'] = "NA"
                self.results['Tension Capacity'] = "NA"
                self.results['Shear X DCR'] = "NG"
                self.results['Shear Y DCR'] = "NG"
                self.results['Shear DCR'] = "NG"
                self.results['Tension DCR'] = "NG"
                self.results['OK'] = False
                self.DCR = np.inf
            else:
                self.results['Shear X Capacity'] = filtered_dfx['shear'].values[0]
                self.results['Shear Y Capacity'] = filtered_dfy['shear'].values[0]
                self.results['Tension Capacity'] = filtered_dfx['tension'].values[0]
                self.results['Shear X DCR'] = self.results['Shear X Demand'] / self.results['Shear X Capacity']
                self.results['Shear Y DCR'] = self.results['Shear Y Demand'] / self.results['Shear Y Capacity']
                self.results['Shear DCR'] = (self.results['Shear X DCR']**2+self.results['Shear Y DCR']**2)**0.5
                self.results['Tension DCR'] = self.results['Tension Demand'] / self.results['Tension Capacity']
                self.results['OK'] = self.results['Shear DCR'] < 1 and self.results['Tension DCR'] < 1
                self.DCR = max(self.results['Tension DCR'], self.results['Shear DCR'])
        else:
            self.results['Shear X Capacity'] = 'NA'
            self.results['Shear Y Capacity'] = 'NA'
            self.results['Tension Capacity'] = 'NA'
            self.results['Shear X DCR'] = 'NG'
            self.results['Shear Y DCR'] = 'NG'
            self.results['Shear DCR'] = "NG"
            self.results['Tension DCR'] = 'NG'
            self.results['OK'] = False
            self.DCR = np.inf

    def reset_results(self):
        self.results = {}

    def max_dcr(self):
        """For use in comparing dcrs of multiple SMSAnchors objects.
        Returns the max of shear or tension DCR, or inf if either is NG."""
        if self.results == {}:
            return np.inf
        shear_x = np.inf if self.results['Shear X DCR'] == 'NG' else self.results['Shear X DCR']
        shear_y = np.inf if self.results['Shear Y DCR'] == 'NG' else self.results['Shear Y DCR']
        tension = np.inf if self.results['Tension DCR'] == 'NG' else self.results['Tension DCR']
        return max(shear_x, shear_y, tension)


class WoodFastener:
    EDGE_DIST_REQS = pd.dataframe({
        'loading_dir':['perpendicular', 'parallel_compression','parallel_tension', 'parallel_tension'],
        'wood_class':[None,None,'softwood','hardwood'],
        'minimum_for_05': [2, 2, 3.5, 2.5],
        'minimum_for_1': [4, 4, 7 ,5]
    })
    def __init__(self, wall_data=None, xy_anchors=None, backing_type=None, df_sms=None, condition_x = None,
                 condition_y = None):
        self.fastener_type = fastener_type
        self.wood_class = None # hardwood or softwood

        self.C_delta = None
        self.C_eg
        self.C_di
        self.C_tn


        pass

    def set_fastener_properties(self):

        # Steel Bearing
        # self.Fes = <something> todo

        ''' COMPUTED PROPERITES'''
        theta_rad = np.radians(self.theta)

        # Wood Bearing (NDS 12.3.4)
        if self.D < 0.25:
            self.Fe = 16600 * self.G**1.84
        else:
            self.Fe_parallel = 11200*self.G
            self.Fe_perp = 6100*G**1.45/(self.D**0.5)
            self.Fem = self.Fe_parallel * self.Fe_perp / (self.Fe_parallel * np.sin(theta_rad)**2 +
                                                          self.Fe_perp * np.sin(theta_rad)**2)

        # K_theta from table 12.3.1B
        self.K_theta = 1 + 0.25 * (self.theta/90)

        #K_D from Table 12.3.1B
        if self.D <= 0.17:
            self.K_D = 2.2
        elif 0.17 < self.D < 0.25:
            self.K_D = 10 * self.D + 0.5
        else:
            self.K_D = np.nan

    def reference_lateral_design_value(self):
        D = self.D
        # Fyb = # Get this from fastener data
        K_D = self.K_D
        K_theta = self.K_theta

        if D < 0.25:
            Rd = [K_D]*6
        elif (D > 0.25) and (self.D_root < 0.25):
            Rd = [K_D * K_theta]*6
        elif 0.25 < D < 1:
            Rd = [4*K_theta]*2 + [3.6*K_theta] + [3.2*K_theta]*3
        else:
            Rd = [np.inf]*6
            raise Exception("Fastener diameter greater than 1\" not permitted")

        Rt = lm/ls
        Re = Fem/Fes

        k1 = (Re + 2*Re**2*(1 + Rt + Rt**2) + Rt**2*Re**3)**0.5 - Re*(1+Rt) / (1+Re)
        k2 = -1 + (2 * (1 + Re) + 2 * Fyb * (1 + 2 * Re) * D ** 2 / (3 * Fem * lm ** 2)) ** 0.5
        k3 = -1 + (2 * (1 + Re) / Re + 2 * Fyb * (2 + Re) * D ** 2 / (3 * Fem * ls ** 2)) ** 0.5

        # Yield Limit Equations (NDS 12.3.1)
        self.yield_modes = {'Im': D * lm * Fem / Rd[0],
                            'Is': D * ls * Fes / Rd[1],
                            'II': k1 *  D * ls * Fes / Rd[2],
                            'IIIm': k2 * D * lm * Fem / ( (1 + 2 * Re) * Rd[3]),
                            'IIIs': k3 * D * ls * Fem / ( (2 + Re) + Rd[4]),
                            'IV': D**2 / Rd[5] * (2 * Fem * Fyb / (3 * (1 + Re)))**0.5}

        self.z = min([val for key, val in self.yield_modes.items()])

        self.z_prime =

    def reference_withdrawal_design_value(self):
        if self.fastener_type == 'Lag Screw':
            self.w = 1800*self.G**(3/2)*self.D**(3/4)
        elif self.fastener_type == 'Wood Screw':
            self.w = 1380 * self.G ** (5 / 2) ** self.D
        else:
            raise Exception(f'Wood fastener type {self.fastener_type} not supported.')

    def get_loading_dir(self):
        """ Determines the direction of loading relative to the wood grain"""
        #todo: add some logic here
        self.grain_angle

        if
        self.loading_condition = 'perpendicular'
        self.theta = 90


    def adjustment_factors(self):

        # Geometry Factor
        edge_dist_reqs = {'perpendicular': (2*self.D, 4*self.D),
                          'parallel_compression': (2*self.D, 4*self.D),
                          'parallel_tension_hardwood': (2.5*self.D, 5*self.D),
                          'parallel_tension_softwood': ()}

        if self.D < 0.25:
            self.C_delta = 1.0
        else:
            loading_dir = TEST
            min_edge_05_coef, min_edge_1_coef = WoodFastener.EDGE_DIST_REQS[loading_dir]


    def check_anchors(self):




        self.w_prime
        self.z_prime
        self.alpha

        self.z_alpha_prime = self.w_prime * p * z_prime / (self.w_prime*self.p*np.cos(self.alpha)**2+
                                                           self.z_prime*np.sin(self.alpha)**2)




        # self.DCR =

class BraceElement:
    def __init__(self, xyz_i, xyz_j):
        """ Assumes the brace element attaches the equipment object to a hardware object
        (for now, floor plate element)"""

        self.xyz_i = xyz_i
        self.xyz_j = xyz_j
        self.ks = None
        self.tension_only = False
        self.bracket_id = None
        self.capacity_to_equipment = None
        self.brace_capacity = None
        self.capacity_to_backing = None
        self.M = None
        self.k_element = None
        self.brace_force = None
        self.tension_dcr = None
        self.capacity_method = None

    def set_brace_properties(self, bracket_data):
        self.ks = bracket_data['bracket_stiffness']
        self.tension_only = bool(bracket_data['tension_only'])
        self.bracket_id = bracket_data['bracket_id']
        self.capacity_to_equipment = bracket_data['capacity_to_equipment']
        self.brace_capacity = bracket_data['bracket_capacity']
        self.capacity_to_backing = bracket_data['capacity_to_backing']
        self.capacity_method = bracket_data['capacity_method']

    def get_element_deformation(self, u):
        return self.M @ u

    def get_element_stiffness_matrix(self, u, initial=False):
        if self.tension_only and not initial and (self.get_element_deformation(u) < 0):
            return np.zeros_like(self.k_element)
        else:
            return self.k_element

    def get_brace_force(self, u):
        force = self.ks * self.M @ u
        if self.tension_only:
            self.brace_force = max([force, 0])
        else:
            self.brace_force = force

    def check_brace(self, u, asd_lrfd_ratio=1):
        self.get_brace_force(u)
        if self.capacity_method == 'ASD':
            brace_force = self.brace_force * asd_lrfd_ratio
        else:
            brace_force = self.brace_force

        capacities = [self.capacity_to_equipment, self.brace_capacity, self.capacity_to_backing]
        governing_capacity = min([item for item in capacities if isinstance(item, (int, float))], default=None)
        self.tension_dcr = brace_force / governing_capacity if governing_capacity else 'OK by inspection'


class BaseStrap(BraceElement):
    def __init__(self, xyz_i, xyz_j, base_plate):
        super().__init__(xyz_i, xyz_j)
        self.base_plate = base_plate

    def pre_compute_matrices(self):
        (dx, dy, dz) = (j - i for j, i in zip(self.xyz_j, self.xyz_i))
        L = (dx ** 2 + dy ** 2 + dz ** 2) ** 0.5

        A = np.array([-dx / L, -dy / L, -dz / L, dx / L, dy / L, dz / L])

        (xi, yi, zi) = self.xyz_i
        (xj, yj, zj) = self.xyz_j
        x0 = self.base_plate.x0
        y0 = self.base_plate.y0
        z0 = self.base_plate.z0

        ci = np.array([[1, 0, 0, 0, zi, -yi],
                       [0, 1, 0, -zi, 0, xi],
                       [0, 0, 1, yi, -xi, 0]])
        cj = np.array([[1, 0, 0, 0, zj - z0, -(yj - y0)],
                       [0, 1, 0, -(zj - z0), 0, (xj - x0)],
                       [0, 0, 1, (yj - y0), -(xj - x0), 0]])

        local_constraints = np.block([
            [ci, np.zeros((3, 6))],
            [np.zeros((3, 6)), cj]])

        c_element = self.base_plate.C
        c_equip = np.eye(*c_element.shape)

        global_constraints = np.vstack((c_equip, c_element))

        self.M = A @ local_constraints @ global_constraints
        self.k_element = self.ks * np.outer(self.M.T, self.M)
        return
